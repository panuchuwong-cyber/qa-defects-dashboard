"""Create professional Excel template for QA Defects data"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

output_path = "/Users/panuchuwong/Desktop/Kanom/QA_Defects_Dashboard/QA_Defects_Template.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Defects Data"

# ============================================================
# STYLES
# ============================================================
HEADER_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=12, name="Arial")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(size=11, name="Arial")

ALT_FILL = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="888888"),
    right=Side(style="thin", color="888888"),
    top=Side(style="thin", color="888888"),
    bottom=Side(style="thin", color="888888"),
)

THICK_BOTTOM = Border(
    left=Side(style="thin", color="888888"),
    right=Side(style="thin", color="888888"),
    top=Side(style="thin", color="888888"),
    bottom=Side(style="medium", color="000000"),
)

# ============================================================
# TITLE ROW
# ============================================================
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "⚡ 3K BATTERY - QA DEFECTS DATA ENTRY"
title_cell.font = Font(bold=True, color="FFD700", size=16, name="Arial")
title_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Subtitle row
ws.merge_cells("A2:H2")
sub_cell = ws["A2"]
sub_cell.value = "📋 Fill in new defect records below. Dashboard auto-syncs every 30 seconds."
sub_cell.font = Font(italic=True, color="666666", size=10)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Empty row
ws.row_dimensions[3].height = 8

# ============================================================
# HEADER ROW (row 4)
# ============================================================
headers = ["Date", "Supplier", "Group Part", "Problem Mode",
           "Part Name", "Part No", "Qty", "Comment"]
HEADER_ROW = 4

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THICK_BOTTOM

ws.row_dimensions[HEADER_ROW].height = 28

# ============================================================
# DATA VALIDATIONS (dropdowns)
# ============================================================
# Group Part dropdown
group_dv = DataValidation(
    type="list",
    formula1='"ELECTRIC & ELEC.,PACKING,PIPING,PLASTIC,PRINTING,RAW MATERIAL,RUBBER,SHEET METAL,FOAM,SEALING,OTHERS"',
    allow_blank=True,
    showDropDown=False  # False = SHOW dropdown arrow
)
group_dv.error = "Please select a valid Group Part"
group_dv.errorTitle = "Invalid Group Part"
group_dv.prompt = "Select Group Part from dropdown"
group_dv.promptTitle = "Group Part"
ws.add_data_validation(group_dv)
group_dv.add(f"C5:C10000")

# Problem Mode dropdown
mode_dv = DataValidation(
    type="list",
    formula1='"APPEARANCE NG,PART MISTAKE,DIMENSION NG,FUNCTION NG,LEAK"',
    allow_blank=True,
    showDropDown=False
)
mode_dv.error = "Please select a valid Problem Mode"
mode_dv.errorTitle = "Invalid Problem Mode"
mode_dv.prompt = "Select Problem Mode from dropdown"
mode_dv.promptTitle = "Problem Mode"
ws.add_data_validation(mode_dv)
mode_dv.add(f"D5:D10000")

# Date format validation
date_dv = DataValidation(
    type="date",
    operator="between",
    formula1="2020-01-01",
    formula2="2030-12-31",
    allow_blank=True
)
date_dv.error = "Please enter a valid date (YYYY-MM-DD)"
date_dv.errorTitle = "Invalid Date"
date_dv.prompt = "Enter date as YYYY-MM-DD (e.g., 2026-08-31)"
date_dv.promptTitle = "Date"
ws.add_data_validation(date_dv)
date_dv.add(f"A5:A10000")

# Qty must be positive integer
qty_dv = DataValidation(
    type="whole",
    operator="greaterThan",
    formula1=0,
    allow_blank=True
)
qty_dv.error = "Quantity must be a positive number"
qty_dv.errorTitle = "Invalid Quantity"
qty_dv.prompt = "Enter quantity (positive integer)"
qty_dv.promptTitle = "Quantity"
ws.add_data_validation(qty_dv)
qty_dv.add(f"G5:G10000")

# ============================================================
# SAMPLE DATA ROWS (5 examples for users to see format)
# ============================================================
from datetime import datetime, timedelta

sample_data = [
    [datetime(2026, 8, 31), "NISSEN CHEMITEC", "PIPING", "APPEARANCE NG",
     "ELBOW PIPE 1/2", "1P095004-1 K", 120, "Surface scratch on outer diameter"],
    [datetime(2026, 9, 1), "PARADISE", "SHEET METAL", "APPEARANCE NG",
     "INSTALLATION PLATE", "3P811368-1 A", 51, "Dent mark on corner"],
    [datetime(2026, 9, 2), "KSV", "ELECTRIC & ELEC.", "LEAK",
     "CAPACITOR 450V", "VE101-22A", 12, "Insulation damage detected"],
    [datetime(2026, 9, 3), "AKUSAN", "ELECTRIC & ELEC.", "APPEARANCE NG",
     "CONNECTOR 2PIN", "AC2019-A", 8, "Pin misalignment"],
    [datetime(2026, 9, 4), "BTD", "PLASTIC", "PART MISTAKE",
     "COVER HOUSING", "4P022001-2 B", 25, "Wrong color (Black vs Grey)"],
]

for i, row_data in enumerate(sample_data):
    row_num = HEADER_ROW + 1 + i  # Start from row 5
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER
        if i % 2 == 1:
            cell.fill = ALT_FILL

        # Format Date column as YYYY-MM-DD
        if col_idx == 1 and isinstance(value, datetime):
            cell.number_format = "YYYY-MM-DD"

# ============================================================
# COLUMN WIDTHS
# ============================================================
col_widths = {
    "A": 13,   # Date
    "B": 20,   # Supplier
    "C": 18,   # Group Part
    "D": 18,   # Problem Mode
    "E": 24,   # Part Name
    "F": 16,   # Part No
    "G": 10,   # Qty
    "H": 40,   # Comment
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Freeze panes (freeze header rows)
ws.freeze_panes = "A5"

# ============================================================
# INSTRUCTIONS SHEET
# ============================================================
ws2 = wb.create_sheet("📖 Instructions", 0)  # Insert as first sheet

instructions = [
    ("⚡ 3K BATTERY - QA DEFects Data Entry Template", True, 16),
    ("", False, 11),
    ("📋 HOW TO USE", True, 14),
    ("", False, 11),
    ("1. Fill in defect records on the 'Defects Data' sheet", False, 11),
    ("2. Required columns: Date, Supplier, Group Part, Problem Mode, Part No, Qty", False, 11),
    ("3. Optional: Part Name, Comment", False, 11),
    ("4. Use dropdowns for Group Part and Problem Mode", False, 11),
    ("5. Date format: YYYY-MM-DD (e.g., 2026-08-31)", False, 11),
    ("6. Qty must be a positive integer", False, 11),
    ("", False, 11),
    ("🚀 HOW TO SYNC TO DASHBOARD", True, 14),
    ("", False, 11),
    ("Method 1: One-click Sync (Mac) ⭐ RECOMMENDED", True, 12),
    ("   • Save this file after editing", False, 11),
    ("   • Double-click 'sync_to_dashboard.command' on Desktop", False, 11),
    ("   • Wait 3 seconds — dashboard updates automatically", False, 11),
    ("", False, 11),
    ("Method 2: Manual Git Push", True, 12),
    ("   • Save this file after editing", False, 11),
    ("   • Open Terminal → run: ./sync_data.sh", False, 11),
    ("   • Or send the file to Kanom via Telegram", False, 11),
    ("", False, 11),
    ("Method 3: Upload via Dashboard", True, 12),
    ("   • Save this file after editing", False, 11),
    ("   • Login to dashboard → Sidebar → Upload Excel", False, 11),
    ("   • Select this file → Click Upload", False, 11),
    ("", False, 11),
    ("📊 DASHBOARD URL", True, 14),
    ("   https://defects-test.streamlit.app", False, 11),
    ("   Password: TEST@2026", False, 11),
    ("", False, 11),
    ("❓ SUPPORT", True, 14),
    ("   Telegram: @KanomBot", False, 11),
    ("   Response time: < 5 minutes during work hours", False, 11),
]

current_row = 1
for text, is_bold, font_size in instructions:
    cell = ws2.cell(row=current_row, column=1, value=text)
    cell.font = Font(bold=is_bold, size=font_size, name="Arial",
                     color="FFD700" if is_bold and font_size >= 14 else "000000")
    if is_bold and font_size >= 14:
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.font = Font(bold=True, size=font_size, name="Arial", color="FFD700")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    current_row += 1

ws2.column_dimensions["A"].width = 80

# ============================================================
# REFERENCE SHEET (list of valid values for dropdowns)
# ============================================================
ws3 = wb.create_sheet("📚 Reference")

# Group Parts
ws3["A1"] = "Group Parts"
ws3["A1"].font = Font(bold=True, size=14, color="FFD700")
ws3["A1"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

group_parts = [
    "ELECTRIC & ELEC.", "PACKING", "PIPING", "PLASTIC", "PRINTING",
    "RAW MATERIAL", "RUBBER", "SHEET METAL", "FOAM", "SEALING", "OTHERS"
]
for i, g in enumerate(group_parts, start=2):
    ws3.cell(row=i, column=1, value=g).font = Font(size=11)

# Problem Modes
ws3["C1"] = "Problem Modes"
ws3["C1"].font = Font(bold=True, size=14, color="FFD700")
ws3["C1"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

problem_modes = [
    "APPEARANCE NG", "PART MISTAKE", "DIMENSION NG", "FUNCTION NG", "LEAK"
]
for i, m in enumerate(problem_modes, start=2):
    ws3.cell(row=i, column=3, value=m).font = Font(size=11)

# Suppliers (existing)
ws3["E1"] = "Suppliers"
ws3["E1"].font = Font(bold=True, size=14, color="FFD700")
ws3["E1"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

suppliers = [
    "NISSEN CHEMITEC", "PARADISE", "BTD", "C.G.", "TTS PLASTIC",
    "KSV", "AKUSAN", "SAMBO", "DEM", "APT (Thailand)",
    "SUPERFAST", "SAM NEO", "TECHNO ASSOCIATE", "[ADD NEW...]"
]
for i, s in enumerate(suppliers, start=2):
    ws3.cell(row=i, column=5, value=s).font = Font(size=11)

ws3.column_dimensions["A"].width = 25
ws3.column_dimensions["C"].width = 25
ws3.column_dimensions["E"].width = 25

# ============================================================
# SAVE
# ============================================================
# Set Instructions as active sheet
wb.active = ws2

wb.save(output_path)
print(f"✅ Excel template created: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Sample data rows: 5")
print(f"   Data validations: Group Part, Problem Mode, Date, Qty")
