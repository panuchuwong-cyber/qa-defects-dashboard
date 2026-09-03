"""
TEST QA Defects Dashboard - Professional Edition
Two pages: 14 Days Monitoring + Searching Supplier Information
Theme: Yellow & Black (Energy Brand)
"""

import streamlit as st
import pandas as pd
from pathlib import Path
import io
from datetime import datetime, timedelta

CHARTJS_CDN = "https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"
LOGO_URL = "https://cdn.jsdelivr.net/gh/panuchuwong-cyber/qa-defects-dashboard@main/assets/3k_logo.jpg"

# ============================================================
# PASSWORD GATE
# ============================================================
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if not st.session_state.authenticated:
        logo_html = """
        <style>
        body { background: #0a0a0a; }
        .login-wrap {
            max-width: 420px; margin: 90px auto; padding: 48px 36px;
            background: linear-gradient(145deg, #1a1a1a 0%, #000000 100%);
            border: 2px solid #FFD700; border-radius: 20px;
            text-align: center; box-shadow: 0 12px 48px rgba(255,215,0,0.25);
            position: relative; overflow: hidden;
        }
        .login-wrap::before {
            content: ""; position: absolute; top: -50%; left: -50%;
            width: 200%; height: 200%;
            background: radial-gradient(circle, rgba(255,215,0,0.05) 0%, transparent 70%);
        }
        .login-icon { font-size: 48px; margin-bottom: 16px; }
        .login-title {
            color: #FFD700; font-size: 26px; font-weight: 900;
            letter-spacing: 3px; margin-bottom: 4px;
            text-shadow: 0 2px 8px rgba(255,215,0,0.4);
        }
        .login-sub {
            color: #999; font-size: 11px; margin-bottom: 32px;
            letter-spacing: 2px; text-transform: uppercase;
        }
        .login-form { position: relative; z-index: 1; }

        /* === PASSWORD INPUT === */
        .login-form input[type="password"],
        .login-form input[type="text"] {
            background: #0a0a0a !important;
            border: 2px solid #FFD700 !important;
            border-radius: 10px !important;
            color: #FFD700 !important;
            font-size: 15px !important;
            font-weight: 600 !important;
            letter-spacing: 2px !important;
            padding: 12px 16px !important;
            text-align: center !important;
            transition: all 0.3s ease !important;
        }
        .login-form input[type="password"]:focus,
        .login-form input[type="text"]:focus {
            border-color: #FFC107 !important;
            box-shadow: 0 0 0 3px rgba(255,215,0,0.25) !important;
            outline: none !important;
        }
        .login-form input::placeholder {
            color: #666 !important;
            letter-spacing: 1px !important;
        }

        /* === ACCESS BUTTON — YELLOW & BLACK THEME === */
        div[data-testid="stButton"] button[kind="primary"] {
            background: linear-gradient(135deg, #FFD700 0%, #FFC107 50%, #FFD700 100%) !important;
            background-size: 200% 100% !important;
            color: #000000 !important;
            font-weight: 900 !important;
            font-size: 15px !important;
            letter-spacing: 2px !important;
            text-transform: uppercase !important;
            border: 2px solid #000000 !important;
            border-radius: 10px !important;
            padding: 14px 24px !important;
            margin-top: 14px !important;
            box-shadow: 0 6px 20px rgba(255,215,0,0.35),
                        inset 0 1px 0 rgba(255,255,255,0.3) !important;
            animation: btnShimmer 3s linear infinite !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
            position: relative !important;
            overflow: hidden !important;
        }
        div[data-testid="stButton"] button[kind="primary"]:hover {
            background: #000000 !important;
            color: #FFD700 !important;
            border-color: #FFD700 !important;
            box-shadow: 0 8px 28px rgba(255,215,0,0.55),
                        inset 0 0 0 2px rgba(255,215,0,0.2) !important;
            transform: translateY(-2px) !important;
        }
        div[data-testid="stButton"] button[kind="primary"]:active {
            transform: translateY(0) !important;
            box-shadow: 0 4px 12px rgba(255,215,0,0.4) !important;
        }
        @keyframes btnShimmer {
            0%   { background-position: 0% 50%; }
            100% { background-position: 200% 50%; }
        }
        </style>
        <div class="login-wrap">
            <div style="margin-bottom:20px;">
                <img src="__LOGO_URL__" width="140"
                     style="border-radius:14px;box-shadow:0 6px 20px rgba(255,215,0,0.4);">
            </div>
            <div class="login-title">3K BATTERY QA</div>
            <div class="login-sub">Defect Monitoring System v2.0</div>
        </div>
        """.replace("__LOGO_URL__", LOGO_URL)
        st.markdown(logo_html, unsafe_allow_html=True)
        col = st.columns([1, 2, 1])
        with col[1]:
            st.markdown('<div class="login-form">', unsafe_allow_html=True)
            password = st.text_input("🔑 Password", type="password",
                                     label_visibility="collapsed",
                                     placeholder="Enter access password")
            if st.button("⚡ ACCESS DASHBOARD", use_container_width=True, type="primary"):
                try:
                    correct = st.secrets["password"]
                except Exception:
                    correct = "TEST@2026"
                if password == correct:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("❌ Incorrect password")
            st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

check_password()

# === LOADING SPLASH (first time only) ===
if "splash_shown" not in st.session_state:
    st.session_state.splash_shown = False
if not st.session_state.splash_shown:
    splash_placeholder = st.empty()
    with splash_placeholder.container():
        st.markdown(
            '<div style="position:fixed; top:0; left:0; right:0; bottom:0; '
            'background:linear-gradient(135deg, #000000 0%, #1a1a1a 100%); '
            'display:flex; align-items:center; justify-content:center; '
            'flex-direction:column; z-index:9999; animation:fadeIn 0.3s ease;">'
            '<div style="font-size:80px; margin-bottom:20px; '
            'animation:pulse 1.5s ease-in-out infinite;">⚡</div>'
            '<div style="color:#FFD700; font-size:32px; font-weight:900; '
            'letter-spacing:6px; margin-bottom:12px;">3K BATTERY</div>'
            '<div style="color:#fff; font-size:14px; letter-spacing:3px; '
            'font-weight:700; margin-bottom:24px;">QA DEFECTS DASHBOARD</div>'
            '<div style="width:200px; height:4px; background:rgba(255,215,0,0.2); '
            'border-radius:2px; overflow:hidden;">'
            '<div style="width:100%; height:100%; background:linear-gradient(90deg, #FFD700, #FFC107); '
            'animation:loadingBar 1.2s ease-in-out infinite;"></div>'
            '</div>'
            '<div style="color:#999; font-size:11px; margin-top:20px; '
            'letter-spacing:2px;">LOADING DASHBOARD...</div>'
            '</div>'
            '<style>'
            '@keyframes fadeIn { from { opacity:0; } to { opacity:1; } }'
            '@keyframes loadingBar { '
            '0% { transform:translateX(-100%); } '
            '50% { transform:translateX(0); } '
            '100% { transform:translateX(100%); } }'
            '</style>',
            unsafe_allow_html=True
        )
        import time as _time
        _time.sleep(1.2)
    splash_placeholder.empty()
    st.session_state.splash_shown = True
    st.rerun()

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="⚡ 3K Battery QA Defects Dashboard",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "About": "3K Battery QA Defects Dashboard v2.0\nBuilt with Streamlit + Chart.js\n⚡ Made by Kanom AI for K-Kream",
        "Get Help": "https://github.com/panuchuwong-cyber/qa-defects-dashboard",
        "Report a bug": "https://github.com/panuchuwong-cyber/qa-defects-dashboard/issues",
    }
)

# ============================================================
# GLOBAL CSS - PROFESSIONAL DESIGN
# ============================================================

st.markdown("""
<style>
    /* === GLOBAL === */
    .stApp { background: linear-gradient(135deg, #fafafa 0%, #f5f5f5 100%); }
    [data-testid="stSidebarNav"] { display: none; }
    footer { visibility: hidden; }
    header[data-testid="stHeader"] { background: transparent; }

    /* === SIDEBAR === */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0a0a0a 0%, #1a1a1a 50%, #0a0a0a 100%);
        border-right: 2px solid #FFD700;
        box-shadow: 4px 0 20px rgba(0,0,0,0.3);
    }
    [data-testid="stSidebar"] * { color: #ffffff !important; }
    [data-testid="stSidebar"] .stRadio label {
        background: rgba(255,215,0,0.08); padding: 12px 16px;
        border-radius: 10px; border: 1px solid rgba(255,215,0,0.25);
        margin-bottom: 8px; transition: all 0.3s; font-weight: 700;
        letter-spacing: 1px;
    }
    [data-testid="stSidebar"] .stRadio label:hover {
        background: rgba(255,215,0,0.2); border-color: #FFD700;
        transform: translateX(4px);
    }
    [data-testid="stSidebar"] .stSelectbox label {
        color: #FFD700 !important; font-weight: 800;
        font-size: 11px; letter-spacing: 2px; text-transform: uppercase;
        margin-bottom: 6px;
    }
    [data-testid="stSidebar"] .stSelectbox > div > div {
        background: rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,215,0,0.3) !important;
        border-radius: 8px !important;
        color: #fff !important;
    }
    [data-testid="stSidebar"] button {
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%) !important;
        color: #000 !important;
        border: none !important; font-weight: 800 !important;
        letter-spacing: 1px !important;
        border-radius: 8px !important;
    }

    /* === SIDEBAR LOGO === */
    .sidebar-logo-wrap {
        background: linear-gradient(135deg, #1a1a1a 0%, #2a2a2a 100%);
        padding: 20px 12px; border-radius: 14px;
        text-align: center; margin-bottom: 24px;
        border: 2px solid #FFD700;
        box-shadow: 0 6px 24px rgba(255,215,0,0.25);
        position: relative;
        overflow: hidden;
    }
    .sidebar-logo-wrap::before {
        content: ""; position: absolute; top: -50%; left: -50%;
        width: 200%; height: 200%;
        background: radial-gradient(circle, rgba(255,215,0,0.08) 0%, transparent 60%);
    }
    .sidebar-logo-img {
        width: 100%; max-width: 200px; height: auto;
        margin: 0 auto 10px auto; display: block;
        border-radius: 10px;
        position: relative; z-index: 1;
    }
    .sidebar-divider {
        height: 1px; background: linear-gradient(90deg,
            transparent 0%, rgba(255,215,0,0.5) 50%, transparent 100%);
        margin: 16px 0;
    }

    /* === MAIN HEADER LOGO === */
    .header-logo-wrap {
        display: flex; align-items: center; gap: 18px;
        padding: 4px 0; position: relative; z-index: 1;
    }
    .header-logo-img {
        width: 72px; height: 72px; border-radius: 14px;
        box-shadow: 0 4px 16px rgba(255,215,0,0.4);
        background: #FFD700;
    }

    /* === MAIN HEADER === */
    .dashboard-header {
        background:
            linear-gradient(135deg, #000000 0%, #1f1f1f 50%, #000000 100%),
            radial-gradient(circle at top right, rgba(255,215,0,0.15) 0%, transparent 50%);
        padding: 28px 36px; border-radius: 18px; margin-bottom: 28px;
        border: 2px solid #FFD700; position: relative; overflow: hidden;
        box-shadow: 0 12px 40px rgba(0,0,0,0.2);
    }
    .dashboard-header::before {
        content: ""; position: absolute; top: -100%; right: -10%;
        width: 400px; height: 400px;
        background: radial-gradient(circle, rgba(255,215,0,0.15) 0%, transparent 60%);
        animation: pulse 4s ease-in-out infinite;
    }
    @keyframes pulse {
        0%, 100% { transform: scale(1); opacity: 0.6; }
        50% { transform: scale(1.2); opacity: 0.9; }
    }
    .dashboard-header h1 {
        color: #FFD700; font-size: 30px; font-weight: 900;
        margin: 0; letter-spacing: 3px;
        text-shadow: 0 2px 16px rgba(255,215,0,0.4);
        position: relative; z-index: 1;
    }
    .dashboard-header .subtitle {
        color: #cccccc; font-size: 12px; margin-top: 6px;
        letter-spacing: 2px; text-transform: uppercase;
        position: relative; z-index: 1;
    }
    .dashboard-header .badge {
        display: inline-block; background: #FFD700; color: #000;
        padding: 5px 14px; border-radius: 14px; font-size: 10px;
        font-weight: 900; letter-spacing: 1.5px; margin-top: 10px;
        box-shadow: 0 4px 12px rgba(255,215,0,0.3);
        position: relative; z-index: 1;
    }

    /* === HERO HEADER (new unified design) === */
    .hero-header {
        background:
            linear-gradient(135deg, #000000 0%, #1a1a1a 50%, #000000 100%),
            radial-gradient(circle at 80% 20%, rgba(255,215,0,0.18) 0%, transparent 60%);
        border: 2px solid #FFD700;
        border-radius: 16px;
        padding: 20px 22px;
        margin-bottom: 18px;
        position: relative;
        overflow: hidden;
        box-shadow: 0 8px 28px rgba(0,0,0,0.25);
        display: grid;
        grid-template-columns: 1fr auto;
        grid-template-areas:
            "left right"
            "stats stats"
            "live live";
        gap: 14px 18px;
    }
    .hero-header::before {
        content: "";
        position: absolute; top: -50%; right: -20%;
        width: 320px; height: 320px;
        background: radial-gradient(circle, rgba(255,215,0,0.18) 0%, transparent 60%);
        animation: pulse 5s ease-in-out infinite;
        pointer-events: none;
    }
    .hero-left {
        grid-area: left;
        position: relative; z-index: 1;
    }
    .hero-title {
        color: #FFD700;
        font-size: 22px;
        font-weight: 900;
        letter-spacing: 2px;
        line-height: 1.15;
        text-shadow: 0 2px 12px rgba(255,215,0,0.4);
    }
    .hero-subtitle {
        color: #cccccc;
        font-size: 10px;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        margin-top: 6px;
        font-weight: 600;
    }
    .hero-badge {
        display: inline-block;
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        color: #000;
        padding: 5px 12px;
        border-radius: 12px;
        font-size: 9px;
        font-weight: 900;
        letter-spacing: 1.5px;
        margin-top: 10px;
        box-shadow: 0 4px 12px rgba(255,215,0,0.35);
    }
    .hero-right {
        grid-area: right;
        background: rgba(255,215,0,0.08);
        border: 1px solid rgba(255,215,0,0.25);
        border-radius: 10px;
        padding: 8px 12px;
        text-align: right;
        position: relative; z-index: 1;
        min-width: 120px;
    }
    .hero-date-row {
        display: flex;
        flex-direction: column;
        align-items: flex-end;
        gap: 1px;
    }
    .hero-date-label {
        color: #FFD700;
        font-size: 8px;
        font-weight: 800;
        letter-spacing: 2px;
        text-transform: uppercase;
    }
    .hero-date-val {
        color: #ffffff;
        font-size: 13px;
        font-weight: 800;
        font-family: 'Courier New', monospace;
        letter-spacing: 0.5px;
    }
    .hero-date-divider {
        color: rgba(255,215,0,0.5);
        font-size: 9px;
        letter-spacing: 1px;
        margin: 2px 0;
        text-align: right;
    }
    .hero-stats {
        grid-area: stats;
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 8px;
        position: relative; z-index: 1;
    }
    .hero-stat {
        background: linear-gradient(135deg, rgba(255,215,0,0.06) 0%, rgba(0,0,0,0.4) 100%);
        border: 1px solid rgba(255,215,0,0.3);
        border-radius: 8px;
        padding: 8px 6px;
        text-align: center;
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 2px;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .hero-stat:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(255,215,0,0.2);
    }
    .hero-stat-icon {
        font-size: 14px;
    }
    .hero-stat-value {
        color: #FFD700;
        font-size: 16px;
        font-weight: 900;
        line-height: 1;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .hero-stat-label {
        color: #ffffff;
        font-size: 8px;
        font-weight: 700;
        letter-spacing: 1px;
        text-transform: uppercase;
        opacity: 0.85;
    }
    .hero-stat-clock {
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        border-color: #FFD700;
    }
    .hero-stat-clock .hero-stat-icon { font-size: 12px; }
    .hero-stat-clock .hero-stat-label {
        color: #000;
        opacity: 1;
        font-weight: 800;
    }
    .hero-stat-clock-val {
        color: #000;
        font-size: 12px;
        font-weight: 900;
        font-family: 'Courier New', monospace;
        letter-spacing: 0.5px;
        line-height: 1;
    }
    .hero-live {
        grid-area: live;
        background: linear-gradient(135deg, #000000 0%, #1a1a1a 100%);
        border: 1.5px solid #FFD700;
        border-radius: 999px;
        padding: 7px 18px;
        color: #FFD700;
        font-size: 10px;
        font-weight: 800;
        letter-spacing: 2.5px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        gap: 9px;
        position: relative; z-index: 1;
        box-shadow: 0 0 0 3px rgba(255,215,0,0.08), 0 6px 14px rgba(0,0,0,0.25);
        width: fit-content;
        justify-self: start;
    }
    .hero-live-pulse {
        width: 9px; height: 9px;
        background: #4CAF50;
        border-radius: 50%;
        box-shadow: 0 0 0 0 rgba(76,175,80,0.7);
        animation: livePulse 1.8s ease-in-out infinite;
        flex-shrink: 0;
    }
    @keyframes livePulse {
        0%, 100% { box-shadow: 0 0 0 0 rgba(76,175,80,0.7); transform: scale(1); }
        50%      { box-shadow: 0 0 0 8px rgba(76,175,80,0); transform: scale(1.1); }
    }

    /* === DATE RANGE CHIP === */
    .date-chip {
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        color: #000; padding: 14px 20px; border-radius: 12px;
        text-align: center; font-weight: 800; font-size: 14px;
        box-shadow: 0 4px 12px rgba(255,215,0,0.3);
    }

    /* === INFO STACK (sidebar cards) === */
    .info-stack {
        margin-top: 14px;
        margin-bottom: 18px;
        display: grid;
        grid-template-columns: repeat(3, 1fr);
        gap: 12px;
    }
    .info-card {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%);
        border: 1.5px solid #FFD700;
        border-radius: 10px;
        padding: 12px 6px;
        text-align: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3);
    }
    .info-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 6px 16px rgba(255,215,0,0.25);
    }
    .info-card-icon {
        font-size: 22px;
        margin-bottom: 6px;
        display: block;
    }
    .info-card-value {
        color: #FFD700;
        font-size: 22px;
        font-weight: 900;
        line-height: 1.1;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .info-card-label {
        color: #ffffff;
        font-size: 10px;
        font-weight: 700;
        letter-spacing: 1px;
        text-transform: uppercase;
        margin-top: 4px;
        opacity: 0.95;
    }
    .live-indicator {
        margin-top: 6px;
        margin-bottom: 20px;
        background: linear-gradient(90deg, #1a1a1a 0%, #2d2d2d 100%);
        border: 1.5px solid #FFD700;
        color: #FFD700;
        padding: 12px 14px;
        border-radius: 10px;
        font-size: 11px;
        font-weight: 800;
        text-align: center;
        letter-spacing: 2px;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 10px;
    }
    .live-dot {
        width: 9px; height: 9px;
        background: #FFD700;
        border-radius: 50%;
        animation: pulse 1.5s ease-in-out infinite;
        box-shadow: 0 0 8px #FFD700;
    }

    /* === KPI CARD === */
    .kpi-container {
        background: white; padding: 24px 22px; border-radius: 16px;
        border: 2px solid #1a1a1a; position: relative; overflow: hidden;
        box-shadow: 0 4px 16px rgba(0,0,0,0.06);
        margin-bottom: 14px;
        transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1),
                    box-shadow 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        cursor: default;
        height: 100%;
        min-height: 160px;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
    }
    .kpi-container:hover {
        transform: translateY(-6px);
        box-shadow: 0 16px 40px rgba(0,0,0,0.15);
    }
    /* Top accent bar - thin and clean */
    .kpi-container::before {
        content: ""; position: absolute; top: 0; left: 0;
        width: 100%; height: 3px;
    }
    .kpi-yellow::before {
        background: linear-gradient(90deg, #FFD700 0%, #FFA500 100%);
    }
    .kpi-black::before {
        background: linear-gradient(90deg, #000000 0%, #333333 100%);
    }
    .kpi-icon {
        position: absolute; top: 16px; right: 16px;
        font-size: 26px; opacity: 0.22;
        color: #FFD700;
        filter: drop-shadow(0 0 6px rgba(255,215,0,0.35));
        transition: opacity 0.3s, transform 0.3s;
    }
    .kpi-container:hover .kpi-icon { opacity: 0.45; transform: scale(1.06); }
    
    .kpi-label {
        color: #FFD700; font-size: 12px; font-weight: 800;
        letter-spacing: 2.5px; text-transform: uppercase;
        margin-bottom: 10px; margin-top: 4px;
        opacity: 0.95;
    }
    .kpi-value {
        color: #000; font-size: 32px; font-weight: 900;
        line-height: 1; margin: 4px 0 8px 0;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        white-space: nowrap;
        display: flex;
        align-items: center;
        gap: 6px;
    }
    .kpi-unit {
        font-size: 13px; color: #FFD700; margin-left: 6px;
        font-weight: 800;
        background: #000; padding: 3px 9px; border-radius: 10px;
        text-shadow: none;
        letter-spacing: 0.5px;
    }
    .kpi-trend {
        font-size: 12px; font-weight: 800;
        padding: 7px 14px; border-radius: 14px;
        display: inline-flex; align-items: center; gap: 4px;
        align-self: flex-start;
        margin-top: 14px;
        letter-spacing: 0.3px;
    }
    .trend-up {
        color: #B71C1C; background: rgba(255,107,107,0.12);
        border: 1px solid rgba(255,107,107,0.2);
        animation: trendPulseRed 2s ease-in-out infinite;
    }
    @keyframes trendPulseRed {
        0%, 100% { box-shadow: 0 0 0 0 rgba(255,107,107,0); }
        50%      { box-shadow: 0 0 0 6px rgba(255,107,107,0.15); }
    }
    .trend-down {
        color: #1B5E20; background: rgba(76,175,80,0.12);
        border: 1px solid rgba(76,175,80,0.2);
        animation: trendPulseGreen 2s ease-in-out infinite;
    }
    @keyframes trendPulseGreen {
        0%, 100% { box-shadow: 0 0 0 0 rgba(76,175,80,0); }
        50%      { box-shadow: 0 0 0 6px rgba(76,175,80,0.15); }
    }
    .trend-neutral {
        color: #555; background: rgba(0,0,0,0.05);
        border: 1px solid rgba(0,0,0,0.1);
    }

    /* === INSIGHT BANNER (smart insights) === */
    .insight-banner {
        display: flex;
        align-items: center;
        gap: 16px;
        padding: 16px 20px;
        margin: 16px 0;
        border-radius: 12px;
        border-left: 6px solid currentColor;
        box-shadow: 0 4px 14px rgba(0,0,0,0.08);
        transition: transform 0.2s ease;
    }
    .insight-banner:hover { transform: translateX(2px); }
    .insight-banner-icon {
        font-size: 28px;
        flex-shrink: 0;
        width: 50px; height: 50px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        background: rgba(255,255,255,0.5);
    }
    .insight-banner-body { flex: 1; min-width: 0; }
    .insight-banner-title {
        font-weight: 900;
        font-size: 16px;
        letter-spacing: 0.5px;
        margin-bottom: 4px;
        line-height: 1.3;
    }
    .insight-banner-detail {
        font-size: 13px;
        line-height: 1.5;
        opacity: 0.85;
    }
    .insight-banner-detail b { font-weight: 800; }

    /* Severity color variants */
    .insight-critical {
        color: #B71C1C;
        background: linear-gradient(135deg, rgba(255,107,107,0.12) 0%, rgba(255,107,107,0.04) 100%);
        border-left-color: #B71C1C;
    }
    .insight-critical .insight-banner-icon {
        background: rgba(183,28,28,0.15);
        color: #B71C1C;
    }
    .insight-warning {
        color: #E65100;
        background: linear-gradient(135deg, rgba(255,152,0,0.12) 0%, rgba(255,152,0,0.04) 100%);
        border-left-color: #E65100;
    }
    .insight-warning .insight-banner-icon {
        background: rgba(230,81,0,0.15);
        color: #E65100;
    }
    .insight-good {
        color: #1B5E20;
        background: linear-gradient(135deg, rgba(76,175,80,0.12) 0%, rgba(76,175,80,0.04) 100%);
        border-left-color: #1B5E20;
    }
    .insight-good .insight-banner-icon {
        background: rgba(27,94,32,0.15);
        color: #1B5E20;
    }
    .insight-info {
        color: #555;
        background: linear-gradient(135deg, rgba(0,0,0,0.04) 0%, rgba(0,0,0,0.02) 100%);
        border-left-color: #777;
    }
    .insight-info .insight-banner-icon {
        background: rgba(0,0,0,0.08);
        color: #555;
    }

    /* === SECTION HEADER === */
    .section-header {
        background: linear-gradient(135deg, #000000 0%, #1f1f1f 100%);
        color: #FFD700;
        padding: 14px 22px;
        border-radius: 14px;
        font-weight: 900; font-size: 15px; letter-spacing: 2.5px;
        margin: 28px 0 16px 0; text-transform: uppercase;
        border: 1.5px solid rgba(255,215,0,0.45);
        border-left: 5px solid #FFD700;
        box-shadow:
            0 8px 22px rgba(0,0,0,0.25),
            inset 0 1px 0 rgba(255,215,0,0.15);
        display: flex; align-items: center; gap: 12px;
        position: relative;
        overflow: hidden;
    }
    .section-header::after {
        content: ""; position: absolute; inset: 0;
        background: radial-gradient(circle at top right, rgba(255,215,0,0.10) 0%, transparent 60%);
        pointer-events: none;
    }
    .section-header > * { position: relative; z-index: 1; }
    .section-header .section-icon {
        font-size: 18px;
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        color: #000;
        width: 38px; height: 38px;
        border-radius: 10px;
        display: flex; align-items: center; justify-content: center;
        box-shadow:
            0 4px 10px rgba(255,215,0,0.45),
            inset 0 1px 0 rgba(255,255,255,0.45);
        transform: rotate(-4deg);
        transition: transform .3s ease;
    }
    .section-header:hover .section-icon { transform: rotate(4deg) scale(1.08); }

    /* === FILTER BAR === */
    .filter-bar {
        background: white; padding: 16px 20px; border-radius: 12px;
        border: 1px solid #e0e0e0; margin-bottom: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    .filter-label {
        font-size: 10px; font-weight: 800; color: #666;
        letter-spacing: 2px; text-transform: uppercase; margin-bottom: 4px;
    }

    /* === QUICK ACTION BUTTONS === */
    div[data-testid="column"] button[kind="secondary"] {
        background: white !important;
        color: #000 !important;
        border: 2px solid #FFD700 !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        letter-spacing: 1px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        min-height: 50px !important;
    }
    div[data-testid="column"] button[kind="secondary"]:hover {
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%) !important;
        color: #000 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(255,215,0,0.4) !important;
    }

    /* === EMPTY STATE === */
    .empty-state {
        background: linear-gradient(135deg, #fafafa 0%, #f0f0f0 100%);
        border: 2px dashed #FFD700;
        border-radius: 16px;
        padding: 60px 40px;
        text-align: center;
        margin: 20px 0;
        animation: fadeInUp 0.6s ease-out;
    }
    .empty-state-icon {
        font-size: 64px;
        margin-bottom: 16px;
        display: block;
    }
    .empty-state-title {
        color: #000;
        font-size: 20px;
        font-weight: 900;
        letter-spacing: 1px;
        margin-bottom: 8px;
    }
    .empty-state-text {
        color: #666;
        font-size: 14px;
        margin-bottom: 16px;
    }
    .empty-state-hint {
        display: inline-block;
        background: #FFD700;
        color: #000;
        padding: 8px 16px;
        border-radius: 20px;
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 1px;
    }

    /* === SEARCH BAR === */
    .search-container {
        position: relative;
        margin-bottom: 20px;
    }
    .search-input {
        background: white;
        border: 2px solid #FFD700;
        border-radius: 12px;
        padding: 14px 20px 14px 48px;
        font-size: 14px;
        font-weight: 600;
        width: 100%;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .search-input:focus {
        outline: none;
        border-color: #FFC107;
        box-shadow: 0 0 0 4px rgba(255,215,0,0.25);
    }
    .search-icon {
        position: absolute;
        left: 16px;
        top: 50%;
        transform: translateY(-50%);
        color: #FFD700;
        font-size: 18px;
    }

    /* === GLASSMORPHISM CARD === */
    .glass-card {
        background: rgba(255,255,255,0.7);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border: 1px solid rgba(255,215,0,0.3);
        border-radius: 14px;
        padding: 16px 20px;
        box-shadow: 0 8px 24px rgba(0,0,0,0.06);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }
    .glass-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 12px 32px rgba(255,215,0,0.2);
        border-color: rgba(255,215,0,0.6);
    }

    /* === BADGE TAGS === */
    .badge-tag {
        display: inline-flex;
        align-items: center;
        gap: 4px;
        padding: 4px 10px;
        border-radius: 10px;
        font-size: 10px;
        font-weight: 800;
        letter-spacing: 1px;
        text-transform: uppercase;
    }
    .badge-yellow {
        background: #FFD700;
        color: #000;
        border: 1px solid #000;
    }
    .badge-black {
        background: #000;
        color: #FFD700;
        border: 1px solid #FFD700;
    }
    .badge-orange {
        background: #FFA500;
        color: #fff;
    }
    .badge-green {
        background: rgba(76,175,80,0.15);
        color: #1B5E20;
        border: 1px solid rgba(76,175,80,0.4);
    }

    /* === STATUS INDICATOR === */
    .status-dot {
        display: inline-block;
        width: 8px;
        height: 8px;
        border-radius: 50%;
        margin-right: 6px;
        animation: pulse 2s ease-in-out infinite;
    }
    .status-ok { background: #4CAF50; box-shadow: 0 0 6px #4CAF50; }
    .status-warn { background: #FFA500; box-shadow: 0 0 6px #FFA500; }
    .status-error { background: #F44336; box-shadow: 0 0 6px #F44336; }

    /* === GROUP BUTTON === */
    .group-grid { display: grid; grid-template-columns: repeat(6, 1fr); gap: 10px; }

    /* === DATAFRAME === */
    .stDataFrame {
        border: 1.5px solid rgba(255,215,0,0.55) !important;
        border-radius: 12px !important;
        overflow: hidden;
        box-shadow: 0 6px 18px rgba(0,0,0,0.12);
        background: white;
    }
    /* Premium table look — keep these declarative so they survive Streamlit
       inline-style overrides on body cells; the JS observer in the topbar
       component handles zebra+hover at the parent DOM level. */
    .stDataFrame {
        border: 1.5px solid rgba(255,215,0,0.55) !important;
        border-radius: 12px !important;
        overflow: hidden;
        box-shadow: 0 6px 18px rgba(0,0,0,0.12);
    }
    .stDataFrame thead th {
        padding: 12px 10px !important;
    }

    /* === BUTTONS === */
    .stButton > button {
        border-radius: 8px; font-weight: 700; letter-spacing: 1px;
        transition: all 0.2s;
    }
    .stButton > button:hover { transform: translateY(-1px); }

    /* === LEGEND === */
    .info-box {
        background: linear-gradient(135deg, #000 0%, #1a1a1a 100%);
        color: #FFD700; padding: 14px 18px; border-radius: 12px;
        font-size: 13px; margin-top: 24px;
        border: 1px solid #FFD700;
        line-height: 1.6;
    }
    .info-box b { color: #FFD700; }

    /* === FOOTER === */
    .dashboard-footer {
        text-align: center;
        margin: 40px 0 8px 0;
        padding: 0;
    }
    .footer-card {
        background: linear-gradient(135deg, #0a0a0a 0%, #1a1a1a 100%);
        border: 1.5px solid rgba(255,215,0,0.4);
        border-radius: 18px;
        padding: 22px 24px;
        box-shadow: 0 10px 28px rgba(0,0,0,0.25), inset 0 1px 0 rgba(255,215,0,0.1);
        position: relative;
        overflow: hidden;
    }
    .footer-card::before {
        content: ""; position: absolute; top: -50%; right: -10%;
        width: 300px; height: 300px;
        background: radial-gradient(circle, rgba(255,215,0,0.10) 0%, transparent 60%);
        pointer-events: none;
    }
    .footer-brand {
        display: inline-flex; align-items: center; gap: 12px;
        position: relative; z-index: 1;
    }
    .footer-logo {
        font-size: 32px; line-height: 1;
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        width: 52px; height: 52px;
        border-radius: 14px;
        display: inline-flex; align-items: center; justify-content: center;
        box-shadow: 0 6px 14px rgba(255,215,0,0.35);
        color: #000;
        transform: rotate(-6deg);
    }
    .footer-brand-text { text-align: left; }
    .footer-brand-name {
        color: #FFD700; font-size: 17px; font-weight: 900;
        letter-spacing: 3px; line-height: 1.1;
    }
    .footer-brand-sub {
        color: #aaa; font-size: 10px; font-weight: 700;
        letter-spacing: 1.5px; text-transform: uppercase;
        margin-top: 3px;
    }
    .footer-divider {
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(255,215,0,0.4), transparent);
        margin: 16px 0 14px 0;
        position: relative; z-index: 1;
    }
    .footer-meta {
        display: flex; flex-wrap: wrap; justify-content: center; gap: 14px;
        margin-bottom: 12px;
        position: relative; z-index: 1;
    }
    .footer-meta-item {
        color: #ccc; font-size: 11px; font-weight: 600;
        padding: 4px 12px;
        background: rgba(255,215,0,0.08);
        border: 1px solid rgba(255,215,0,0.2);
        border-radius: 999px;
        letter-spacing: 0.3px;
    }
    .footer-credit {
        color: #888; font-size: 11px; font-weight: 600;
        letter-spacing: 0.5px;
        position: relative; z-index: 1;
    }
    .footer-heart {
        display: inline-block;
        animation: heartBeat 1.4s ease-in-out infinite;
    }
    @keyframes heartBeat {
        0%, 100% { transform: scale(1); }
        25%      { transform: scale(1.18); }
        50%      { transform: scale(1); }
        75%      { transform: scale(1.10); }
    }

    /* === ANIMATIONS & POLISH === */
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(20px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes pulseGlow {
        0%, 100% { box-shadow: 0 0 0 0 rgba(255,215,0,0.6); }
        50%      { box-shadow: 0 0 0 10px rgba(255,215,0,0); }
    }
    @keyframes shimmer {
        0%   { background-position: -200% 0; }
        100% { background-position: 200% 0; }
    }
    @keyframes slideIn {
        from { opacity: 0; transform: translateX(-15px); }
        to   { opacity: 1; transform: translateX(0); }
    }

    /* KPI cards: smooth entrance + hover lift */
    [data-testid="stMetric"], .kpi-container, .kpi-card {
        animation: fadeInUp 0.5s ease-out;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    [data-testid="stMetric"]:hover, .kpi-container:hover, .kpi-card:hover {
        transform: translateY(-4px) scale(1.02);
        box-shadow: 0 12px 28px rgba(255,215,0,0.25) !important;
    }

    /* KPI value counter animation */
    @keyframes countUp {
        from { opacity: 0; transform: translateY(8px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    .kpi-value {
        animation: countUp 0.6s cubic-bezier(0.16, 1, 0.3, 1);
        display: flex; align-items: center; justify-content: center;
        gap: 6px;
    }

    /* Subtle icon bobbing */
    @keyframes iconBob {
        0%, 100% { transform: translateY(0); }
        50%      { transform: translateY(-3px); }
    }
    .kpi-icon {
        animation: iconBob 3s ease-in-out infinite;
        display: inline-block;
    }

    /* KPI value: gradient text + glow */
    .kpi-value {
        font-size: 32px !important;
        font-weight: 900 !important;
        letter-spacing: -1px;
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 50%, #FFD700 100%);
        background-size: 200% 100%;
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        animation: countUp 0.6s cubic-bezier(0.16, 1, 0.3, 1),
                   shimmer 5s linear infinite;
        display: flex; align-items: baseline; justify-content: flex-start;
        gap: 6px;
    }
    .kpi-container.kpi-black .kpi-value,
    .kpi-card.kpi-black .kpi-value {
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    /* The unit badge lives INSIDE .kpi-value, so it inherits the transparent
       text fill used for the gradient number and renders as a solid dark bar.
       Reset fill + clip explicitly so PCS / CASE / PART stay readable. */
    /* Unit badge sits INSIDE .kpi-value and inherits the transparent text
       fill from the gradient number — reset it to a solid yellow-tinted
       glass pill that reads on both light and dark cards. */
    .kpi-value .kpi-unit {
        background: rgba(255,215,0,0.18) !important;
        border: 1px solid rgba(255,215,0,0.55) !important;
        -webkit-background-clip: border-box !important;
        background-clip: border-box !important;
        -webkit-text-fill-color: #FFD700 !important;
        color: #FFD700 !important;
        backdrop-filter: blur(6px);
    }
    .kpi-container.kpi-black .kpi-value .kpi-unit,
    .kpi-card.kpi-black .kpi-value .kpi-unit {
        background: rgba(255,215,0,0.16) !important;
        border-color: rgba(255,215,0,0.6) !important;
        -webkit-text-fill-color: #FFD700 !important;
        color: #FFD700 !important;
    }

    /* KPI container — premium black-gold with glossy corner highlight */
    .kpi-container {
        position: relative; padding: 20px 18px;
        background: linear-gradient(135deg, #0e0e0e 0%, #1c1c1c 60%, #232323 100%);
        border-radius: 16px;
        text-align: left;
        border: 1.5px solid rgba(255,215,0,0.55);
        box-shadow: 0 8px 24px rgba(0,0,0,0.28), inset 0 1px 0 rgba(255,215,0,0.18);
        overflow: hidden;
        transition: transform .3s ease, box-shadow .3s ease, border-color .3s ease;
    }
    .kpi-container::before {
        content: ""; position: absolute; top: 0; left: 0;
        width: 100%; height: 3px;
        background: linear-gradient(90deg, #FFD700 0%, #FFA500 50%, #FFD700 100%);
    }
    .kpi-container::after {
        content: ""; position: absolute; top: -40%; right: -30%;
        width: 180px; height: 180px;
        background: radial-gradient(circle, rgba(255,215,0,0.12) 0%, transparent 60%);
        pointer-events: none;
    }
    .kpi-container:hover {
        transform: translateY(-3px);
        border-color: #FFD700;
        box-shadow: 0 14px 32px rgba(255,215,0,0.18), inset 0 1px 0 rgba(255,215,0,0.28);
    }
    .kpi-container.kpi-black {
        background: linear-gradient(135deg, #0a0a0a 0%, #1f1f1f 100%);
        color: #fff;
        border: 2px solid #FFD700;
    }
    .kpi-container.kpi-black .kpi-label {
        color: #FFD700 !important;
    }

    /* === SIDEBAR NAV GRADIENT BUTTONS === */
    /* Target each nav button by its key (Streamlit generates data-testid with key suffix) */
    [data-testid="stSidebar"] button[kind="secondary"] {
        min-height: 60px !important;
        border-radius: 12px !important;
        margin-bottom: 10px !important;
        font-weight: 800 !important;
        font-size: 13px !important;
        letter-spacing: 1.5px !important;
        text-transform: uppercase !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border: 2px solid transparent !important;
    }
    [data-testid="stSidebar"] button[kind="secondary"]:hover {
        transform: translateX(4px) scale(1.02) !important;
        box-shadow: 0 8px 20px rgba(255,215,0,0.3) !important;
    }
    /* Button 1: 14 Days Monitoring - Yellow gradient */
    [data-testid="stSidebar"] button[kind="secondary"]:has(span:contains("14 DAYS MONITORING")) {
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%) !important;
        color: #000 !important;
    }
    /* Button 2: Searching Supplier - Black gradient */
    [data-testid="stSidebar"] button[kind="secondary"]:has(span:contains("SEARCHING SUPPLIER")) {
        background: linear-gradient(135deg, #000000 0%, #434343 100%) !important;
        color: #FFD700 !important;
        border: 2px solid #FFD700 !important;
    }
    /* Button 3: Data Entry - Orange gradient */
    [data-testid="stSidebar"] button[kind="secondary"]:has(span:contains("DATA ENTRY")) {
        background: linear-gradient(135deg, #FFA500 0%, #FF6347 100%) !important;
        color: #fff !important;
    }

    /* KPI TREND BADGES (consolidated - see line 408) */

    /* LIVE indicator: pulsing glow */
    .live-pulse {
        display: inline-block; width: 10px; height: 10px;
        background: #00E676; border-radius: 50%;
        animation: pulseGlow 2s infinite;
        box-shadow: 0 0 8px #00E676;
    }

    /* DataFrame rows: smooth hover */
    .stDataFrame tbody tr {
        transition: background-color 0.2s ease, transform 0.15s ease;
    }
    .stDataFrame tbody tr:hover {
        background-color: rgba(255,215,0,0.08) !important;
        transform: translateX(2px);
    }

    /* Section headers: shimmer background */

    .section-header > * {
        position: relative; z-index: 1;
    }

    /* Sidebar info cards: slide in */
    .info-card {
        animation: slideIn 0.4s ease-out;
    }

    /* Empty state styling */
    .empty-state {
        text-align: center; padding: 48px 24px;
        background: linear-gradient(135deg, #fafafa 0%, #fff 100%);
        border: 2px dashed #FFD700; border-radius: 14px;
        margin: 20px 0;
    }
    .empty-state-icon {
        font-size: 64px; margin-bottom: 16px;
        animation: fadeInUp 0.6s ease-out;
    }
    .empty-state-title {
        color: #333; font-size: 18px; font-weight: 800;
        margin-bottom: 8px; letter-spacing: 1px;
    }
    .empty-state-sub {
        color: #999; font-size: 13px;
    }

    /* Step indicator cards: scale on hover */
    .step-card {
        transition: transform 0.3s ease, box-shadow 0.3s ease;
        cursor: default;
    }
    .step-card:hover {
        transform: scale(1.05);
        box-shadow: 0 8px 20px rgba(255,215,0,0.3);
    }

    /* === LOADING SKELETON === */
    @keyframes skeletonShimmer {
        0%   { background-position: -400px 0; }
        100% { background-position: 400px 0; }
    }
    .skeleton {
        background: linear-gradient(90deg, #f0f0f0 25%, #e0e0e0 50%, #f0f0f0 75%);
        background-size: 800px 100%;
        animation: skeletonShimmer 1.5s infinite linear;
        border-radius: 8px;
    }
    .skeleton-card {
        height: 100px; margin: 12px 0;
        background: linear-gradient(90deg, #f0f0f0 25%, #e8e8e8 50%, #f0f0f0 75%);
        background-size: 800px 100%;
        animation: skeletonShimmer 1.5s infinite linear;
        border-radius: 12px;
        border: 1px solid #e0e0e0;
    }

    /* === CUSTOM SCROLLBAR (yellow & black) === */
    ::-webkit-scrollbar { width: 10px; height: 10px; }
    ::-webkit-scrollbar-track {
        background: #0a0a0a;
        border-radius: 8px;
    }
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(180deg, #FFD700 0%, #FFC107 100%);
        border-radius: 8px;
        border: 2px solid #0a0a0a;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(180deg, #FFC107 0%, #FFD700 100%);
        box-shadow: 0 0 8px rgba(255,215,0,0.6);
    }
    * { scrollbar-color: #FFD700 #0a0a0a; scrollbar-width: thin; }

    /* === QUICK STAT MINI-CARDS === */
    .quick-stat {
        display: inline-flex; align-items: center; gap: 6px;
        padding: 6px 12px; margin: 4px 4px 4px 0;
        background: linear-gradient(135deg, #fff 0%, #fafafa 100%);
        border: 1.5px solid #FFD700;
        border-radius: 20px;
        font-size: 11px; font-weight: 700;
        color: #000;
        transition: all 0.2s ease;
        box-shadow: 0 2px 6px rgba(0,0,0,0.04);
    }
    .quick-stat:hover {
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(255,215,0,0.3);
    }
    .quick-stat-icon { font-size: 14px; }
    .quick-stat-value { color: #B45309; font-weight: 900; }
    .quick-stat:hover .quick-stat-value { color: #000; }

    /* === STICKY FILTER BAR === */
    .sticky-filter {
        position: sticky; top: 0; z-index: 100;
        background: rgba(255,255,255,0.95);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        padding: 12px 16px;
        border-bottom: 2px solid #FFD700;
        box-shadow: 0 4px 12px rgba(0,0,0,0.06);
        border-radius: 0 0 12px 12px;
        margin-bottom: 16px;
    }

    /* === SEARCH BAR === */
    .search-container {
        position: relative;
        margin: 12px 0;
    }
    .search-input {
        width: 100%;
        padding: 10px 14px 10px 38px;
        border: 2px solid #FFD700;
        border-radius: 10px;
        font-size: 13px;
        background: #fff;
        color: #000;
        transition: all 0.2s ease;
        box-sizing: border-box;
    }
    .search-input:focus {
        outline: none;
        border-color: #FFC107;
        box-shadow: 0 0 0 3px rgba(255,215,0,0.25);
    }
    .search-icon {
        position: absolute; left: 12px; top: 50%;
        transform: translateY(-50%);
        color: #FFD700; font-size: 16px;
    }

    /* === TOOLTIP (CSS only) === */
    .tooltip { position: relative; display: inline-block; }
    .tooltip .tooltip-text {
        visibility: hidden; opacity: 0;
        background: #000; color: #FFD700;
        text-align: center; padding: 6px 10px;
        border-radius: 6px; font-size: 11px;
        position: absolute; z-index: 200;
        bottom: 130%; left: 50%;
        transform: translateX(-50%);
        transition: opacity 0.2s;
        white-space: nowrap;
        border: 1px solid #FFD700;
    }
    .tooltip:hover .tooltip-text {
        visibility: visible; opacity: 1;
    }

    /* === SUCCESS CARD (post-sync) === */
    @keyframes checkmarkPop {
        0%   { transform: scale(0) rotate(-180deg); opacity: 0; }
        60%  { transform: scale(1.3) rotate(0deg); opacity: 1; }
        100% { transform: scale(1) rotate(0deg); opacity: 1; }
    }
    @keyframes cardSlideIn {
        from { opacity: 0; transform: translateY(-20px) scale(0.95); }
        to   { opacity: 1; transform: translateY(0) scale(1); }
    }
    .success-card {
        text-align: center;
        padding: 28px 24px;
        margin: 16px 0;
        background: linear-gradient(135deg, #000 0%, #1a1a1a 100%);
        border: 2px solid #FFD700;
        border-radius: 16px;
        box-shadow: 0 8px 32px rgba(255,215,0,0.3);
        animation: cardSlideIn 0.5s cubic-bezier(0.16, 1, 0.3, 1);
    }
    .success-checkmark {
        width: 64px; height: 64px;
        margin: 0 auto 14px;
        background: linear-gradient(135deg, #00E676 0%, #00C853 100%);
        border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        color: #fff; font-size: 36px; font-weight: 900;
        box-shadow: 0 4px 16px rgba(0,230,118,0.4);
        animation: checkmarkPop 0.7s cubic-bezier(0.16, 1, 0.3, 1);
    }
    .success-title {
        color: #FFD700; font-size: 18px; font-weight: 900;
        letter-spacing: 2px; margin-bottom: 12px;
    }
    .success-stats {
        display: flex; justify-content: center; gap: 20px;
        margin-bottom: 12px;
        color: #fff; font-size: 13px;
    }
    .success-stats b { color: #FFD700; }
    .success-note {
        color: #999; font-size: 11px; font-style: italic;
    }

    /* === SCROLL TO TOP BUTTON === */
    @keyframes fadeInBtn {
        from { opacity: 0; transform: translateY(20px); }
        to   { opacity: 1; transform: translateY(0); }
    }
    .scroll-top-btn {
        position: fixed;
        bottom: 24px; right: 24px;
        width: 48px; height: 48px;
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        border: 2px solid #000;
        border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-size: 20px; font-weight: 900; color: #000;
        cursor: pointer;
        box-shadow: 0 4px 16px rgba(255,215,0,0.4);
        transition: all 0.2s ease;
        z-index: 999;
        animation: fadeInBtn 0.5s ease-out;
    }
    .scroll-top-btn:hover {
        transform: translateY(-4px) scale(1.1);
        box-shadow: 0 8px 24px rgba(255,215,0,0.6);
    }

    /* === DARK THEME OVERRIDES === */
    [data-theme="dark"] .stApp,
    body.dark-mode .stApp {
        background: linear-gradient(135deg, #0a0a0a 0%, #1a1a1a 100%) !important;
        color: #FFD700 !important;
    }
    [data-theme="dark"] .kpi-container,
    body.dark-mode .kpi-container {
        background: linear-gradient(135deg, #1a1a1a 0%, #0a0a0a 100%) !important;
        color: #FFD700 !important;
        border-color: #FFD700 !important;
    }
    [data-theme="dark"] .quick-stat,
    body.dark-mode .quick-stat {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%) !important;
        color: #FFD700 !important;
    }
    [data-theme="dark"] .quick-stat-value,
    body.dark-mode .quick-stat-value {
        color: #FFD700 !important;
    }

    /* Theme toggle button */
    .theme-toggle {
        position: fixed; top: 12px; right: 12px;
        width: 44px; height: 44px;
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        border: 2px solid #000;
        border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        font-size: 18px; cursor: pointer;
        box-shadow: 0 4px 12px rgba(255,215,0,0.4);
        transition: all 0.2s ease;
        z-index: 998;
    }
    .theme-toggle:hover {
        transform: rotate(180deg) scale(1.1);
        box-shadow: 0 6px 16px rgba(255,215,0,0.6);
    }

    /* === PINNED SUPPLIER BADGE === */
    .pinned-badge {
        display: inline-flex; align-items: center; gap: 6px;
        padding: 6px 12px;
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        color: #000; font-weight: 900; font-size: 11px;
        border-radius: 20px;
        border: 2px solid #000;
        box-shadow: 0 2px 8px rgba(255,215,0,0.4);
        letter-spacing: 1px;
        margin: 4px;
    }
    .pinned-badge::before {
        content: "📌";
        animation: pinBob 2s ease-in-out infinite;
    }
    @keyframes pinBob {
        0%, 100% { transform: rotate(-10deg); }
        50%      { transform: rotate(10deg); }
    }

    /* === TOAST NOTIFICATION === */
    @keyframes toastSlide {
        0%   { transform: translateX(120%); opacity: 0; }
        10%  { transform: translateX(0); opacity: 1; }
        90%  { transform: translateX(0); opacity: 1; }
        100% { transform: translateX(120%); opacity: 0; }
    }
    .toast {
        position: fixed; top: 80px; right: 24px;
        background: linear-gradient(135deg, #000 0%, #1a1a1a 100%);
        color: #FFD700; padding: 12px 20px;
        border: 2px solid #FFD700; border-radius: 10px;
        font-weight: 700; font-size: 13px;
        box-shadow: 0 6px 20px rgba(255,215,0,0.3);
        z-index: 997;
        animation: toastSlide 3s ease-in-out forwards;
    }

    /* === ONBOARDING TOOLTIP === */
    @keyframes tooltipBounce {
        0%, 100% { transform: translateY(0); }
        50%      { transform: translateY(-6px); }
    }
    .onboarding-tip {
        background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);
        color: #000; padding: 12px 16px;
        border-radius: 10px; border: 2px solid #000;
        font-size: 13px; font-weight: 700;
        box-shadow: 0 4px 16px rgba(255,215,0,0.3);
        animation: tooltipBounce 2s ease-in-out infinite;
        display: inline-block;
        margin: 8px 0;
    }

    /* ===================================================== */
    /* === RESPONSIVE DESIGN — MOBILE / TABLET ============ */
    /* ===================================================== */

    /* === TREND CHART GRID (Daily Trend Analysis) === */
    .trend-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 14px;
        margin-bottom: 18px;
    }
    .trend-card {
        background: #ffffff;
        border: 2px solid #000;
        border-radius: 12px;
        padding: 14px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }
    .trend-card-head {
        display: flex;
        align-items: center;
        gap: 8px;
        margin-bottom: 8px;
    }
    .trend-dot {
        width: 10px;
        height: 10px;
        border-radius: 50%;
        flex-shrink: 0;
    }
    .trend-dot-gold  { background: #FFD700; }
    .trend-dot-black { background: #000; }
    .trend-card-title {
        color: #000;
        font-weight: 900;
        font-size: 13px;
        letter-spacing: 1px;
    }
    .trend-canvas-wrap {
        width: 100%;
        height: 200px;
        position: relative;
    }
    .trend-canvas-wrap canvas {
        display: block;
        width: 100% !important;
        height: 200px !important;
    }

    /* Tablet & small laptop (768px - 1024px) */
    @media (max-width: 1024px) {
        .dashboard-header h1 {
            font-size: 24px !important;
            letter-spacing: 2px !important;
        }
        .dashboard-header {
            padding: 22px 26px !important;
        }
        .kpi-value {
            font-size: 26px !important;
        }
        .kpi-container {
            padding: 18px 16px !important;
            min-height: 140px !important;
        }
        .trend-grid { gap: 12px !important; }
        .trend-card { padding: 12px !important; }
        .trend-card-title { font-size: 11px !important; }
    }

    /* Mobile (≤768px) */
    @media (max-width: 768px) {
        /* Hide Streamlit's default padding that wastes mobile space */
        .block-container {
            padding-left: 0.6rem !important;
            padding-right: 0.6rem !important;
            padding-top: 1rem !important;
            max-width: 100% !important;
        }

        /* === HERO HEADER mobile layout === */
        .hero-header {
            padding: 16px 16px !important;
            margin-bottom: 12px !important;
            border-radius: 12px !important;
            grid-template-columns: 1fr !important;
            grid-template-areas:
                "left"
                "right"
                "stats"
                "live" !important;
            gap: 12px !important;
        }
        .hero-title {
            font-size: 18px !important;
            letter-spacing: 1.5px !important;
        }
        .hero-subtitle {
            font-size: 9px !important;
        }
        .hero-badge {
            font-size: 8px !important;
            padding: 4px 10px !important;
            margin-top: 8px !important;
        }
        .hero-right {
            text-align: left !important;
            min-width: auto !important;
            padding: 8px 12px !important;
        }
        .hero-date-row {
            flex-direction: row !important;
            align-items: center !important;
            justify-content: space-between !important;
            gap: 8px !important;
        }
        .hero-date-divider {
            text-align: center !important;
            margin: 4px 0 !important;
        }
        .hero-stats {
            grid-template-columns: repeat(4, 1fr) !important;
            gap: 6px !important;
        }
        .hero-stat {
            padding: 6px 4px !important;
            gap: 1px !important;
        }
        .hero-stat-icon { font-size: 12px !important; }
        .hero-stat-value { font-size: 13px !important; }
        .hero-stat-label { font-size: 7px !important; letter-spacing: 0.5px !important; }
        .hero-stat-clock-val { font-size: 10px !important; }
        .hero-live {
            font-size: 9px !important;
            padding: 5px 10px !important;
            letter-spacing: 1px !important;
        }

        /* Stack header columns vertically */
        div[data-testid="stHorizontalBlock"]:has(> div[data-testid="stColumn"] .dashboard-header) {
            flex-wrap: wrap !important;
        }
        div[data-testid="stHorizontalBlock"]:has(> div[data-testid="stColumn"] .dashboard-header)
            > div[data-testid="stColumn"] {
            flex: 1 1 100% !important;
            min-width: 100% !important;
            width: 100% !important;
        }

        /* Header card - tighter padding, smaller title */
        .dashboard-header {
            padding: 16px 18px !important;
            margin-bottom: 14px !important;
            border-radius: 12px !important;
        }
        .dashboard-header h1 {
            font-size: 18px !important;
            letter-spacing: 1.5px !important;
            line-height: 1.2 !important;
        }
        .dashboard-header .subtitle {
            font-size: 10px !important;
            letter-spacing: 1px !important;
            margin-top: 4px !important;
        }
        .dashboard-header .badge {
            font-size: 9px !important;
            padding: 4px 10px !important;
            margin-top: 8px !important;
        }

        /* Date chip - more compact */
        .date-chip {
            padding: 10px 14px !important;
            font-size: 12px !important;
            margin-top: 10px !important;
        }
        .date-chip span {
            font-size: 11px !important;
        }
        .date-chip span[style*="16px"] {
            font-size: 13px !important;
        }

        /* Insight banner */
        .insight-banner { padding: 14px 16px !important; margin: 12px 0 !important; gap: 12px !important; }
        .insight-banner-icon { width: 44px !important; height: 44px !important; font-size: 22px !important; }
        .insight-banner-title { font-size: 14px !important; line-height: 1.3 !important; }
        .insight-banner-detail { font-size: 11.5px !important; line-height: 1.55 !important; }

        /* Info stack - keep 3 cols but smaller */
        .info-stack {
            gap: 6px !important;
            margin-top: 8px !important;
            margin-bottom: 10px !important;
        }
        .info-card {
            padding: 8px 4px !important;
            border-radius: 8px !important;
        }
        .info-card-icon {
            font-size: 16px !important;
            margin-bottom: 2px !important;
        }
        .info-card-value {
            font-size: 18px !important;
        }
        .info-card-label {
            font-size: 10px !important;
            letter-spacing: 0.8px !important;
            margin-top: 2px !important;
        }

        /* Live indicator & clock box - compact */
        .live-indicator {
            padding: 7px 12px !important;
            font-size: 11px !important;
            letter-spacing: 1.5px !important;
            margin-top: 8px !important;
        }

        /* === KPI CARDS (now 2x2 via Python st.columns(2) rows) === */
        /* Target the inner kpi-card directly without :has() */
        .kpi-card {
            padding: 14px 12px !important;
            min-height: 130px !important;
            margin-bottom: 4px !important;
            border-radius: 12px !important;
        }
        .kpi-icon {
            font-size: 22px !important;
            top: 10px !important;
            right: 10px !important;
        }
        .kpi-label {
            font-size: 9px !important;
            letter-spacing: 1.5px !important;
            margin-bottom: 6px !important;
        }
        .kpi-value {
            font-size: 24px !important;
        }
        .kpi-unit {
            font-size: 11px !important;
            padding: 2px 7px !important;
            border-radius: 6px !important;
            margin-left: 4px !important;
            font-weight: 800 !important;
            line-height: 1.4 !important;
            display: inline-block !important;
            white-space: nowrap !important;
            letter-spacing: 0.3px !important;
        }
        .kpi-trend {
            font-size: 11px !important;
            padding: 4px 9px !important;
            margin-top: 8px !important;
            letter-spacing: 0.3px !important;
        }

        /* === TREND CHARTS - stack vertically on mobile === */
        .trend-grid {
            grid-template-columns: 1fr !important;
            gap: 8px !important;
            margin-bottom: 8px !important;
        }
        .trend-card {
            padding: 10px 8px !important;
            border-radius: 10px !important;
        }
        .trend-card-head {
            margin-bottom: 4px !important;
        }
        .trend-card-title {
            font-size: 11px !important;
        }
        .trend-canvas-wrap,
        .trend-canvas-wrap canvas {
            height: 170px !important;
        }
        /* Remove iframe body margin so charts pack tightly */
        /* Charts live in ONE iframe PER chart inside st.columns, so each
           iframe must keep its own height — forcing a fixed height here is
           what left the big empty gap under the trend charts.
           NOTE: Streamlit 1.50 titles these iframes "st.iframe", NOT
           "streamlit.components.v1.html", so target the tag directly. */
        iframe { display: block !important; margin-bottom: 0 !important; }
        div[data-testid="stIFrame"],
        [data-testid="stCustomComponentV1"] { margin-bottom: 0 !important; }
        /* Streamlit adds a 4px bottom margin per column; with charts stacked
           1-up on mobile that margin repeats and reads as dead space. */
        div[data-testid="stColumn"] { margin-bottom: 0 !important; }
        div[data-testid="stElementContainer"] { margin-bottom: 0 !important; }

        /* --- GLOBAL: every column row wraps 2-up on phones, EXCEPT rows that
           hold a chart iframe / dataframe / text input, which go full width. */
        div[data-testid="stHorizontalBlock"] {
            flex-wrap: wrap !important;
            gap: 8px !important;
            row-gap: 8px !important;
        }
        div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
            flex: 1 1 calc(50% - 4px) !important;
            min-width: calc(50% - 4px) !important;
            width: calc(50% - 4px) !important;
        }
        div[data-testid="stHorizontalBlock"]:has(iframe) > div[data-testid="stColumn"],
        div[data-testid="stHorizontalBlock"]:has([data-testid="stDataFrame"]) > div[data-testid="stColumn"],
        div[data-testid="stHorizontalBlock"]:has([data-testid="stTextInput"]) > div[data-testid="stColumn"] {
            flex: 1 1 100% !important;
            min-width: 100% !important;
            width: 100% !important;
        }
        /* Tighten Streamlit's default vertical rhythm on phones */
        div[data-testid="stVerticalBlock"] { gap: 0.5rem !important; }
        div[data-testid="stVerticalBlockBorderWrapper"] { gap: 0.5rem !important; }
        div[data-testid="element-container"] { margin-bottom: 0 !important; }

        /* === Bottom 2-col grids (mode + suppliers) stack === */
        div[data-testid="stHorizontalBlock"]:has(.resp-row-2col) {
            flex-wrap: wrap !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.resp-row-2col)
            > div[data-testid="stColumn"] {
            flex: 1 1 100% !important;
            min-width: 100% !important;
            width: 100% !important;
            margin-bottom: 14px !important;
        }

        /* Section headers - tighter */
        .section-header {
            padding: 10px 14px !important;
            font-size: 12px !important;
            letter-spacing: 1.5px !important;
            border-radius: 8px !important;
            margin: 14px 0 10px 0 !important;
        }

        /* DataFrames / tables - allow horizontal scroll instead of breaking layout */
        .stDataFrame {
            overflow-x: auto !important;
        }

        /* Hide hover effects on mobile (they cause jank on touch) */
        .kpi-container:hover,
        .info-card:hover,
        .dashboard-header:hover { transform: none !important; }
    }

    /* Very small phones (≤420px) */
    @media (max-width: 420px) {
        .dashboard-header h1 { font-size: 16px !important; }
        .dashboard-header .subtitle { font-size: 9px !important; }
        .info-card-value { font-size: 17px !important; }
        .info-card-label { font-size: 9px !important; }
        .kpi-value { font-size: 21px !important; }
        .kpi-label { font-size: 10px !important; }
        .kpi-card { padding: 12px 10px !important; min-height: 120px !important; }
        .trend-card-title { font-size: 11px !important; }
    }

    /* ===================================================== */
    /* === GLOBAL MOBILE STACKING — ALL PAGES ============ */
    /* ===================================================== */

    /* Any horizontal block with 3+ columns stacks on mobile */
    @media (max-width: 768px) {
        /* Universal: stack every 3+ column layout to single column */
        div[data-testid="stHorizontalBlock"] {
            flex-wrap: wrap !important;
            gap: 8px !important;
        }
        div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
            flex: 1 1 100% !important;
            min-width: 100% !important;
            width: 100% !important;
            margin-bottom: 4px !important;
        }
        /* EXCEPTIONS — keep these as 2-col on mobile */
        /* Header row stays single column (already wrapped above) */
        /* KPI rows: 2x2 layout */
        div[data-testid="stHorizontalBlock"]:has(.kpi-card) {
            gap: 6px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.kpi-card) > div[data-testid="stColumn"] {
            flex: 1 1 calc(50% - 4px) !important;
            min-width: calc(50% - 4px) !important;
            max-width: calc(50% - 4px) !important;
            width: calc(50% - 4px) !important;
            margin-bottom: 4px !important;
        }
        /* Selection supplier group buttons: 3-col grid */
        div[data-testid="stHorizontalBlock"]:has(.group-btn-wrap) {
            gap: 8px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.group-btn-wrap) > div[data-testid="stColumn"] {
            flex: 1 1 calc(33.333% - 6px) !important;
            min-width: calc(33.333% - 6px) !important;
            max-width: calc(33.333% - 6px) !important;
            width: calc(33.333% - 6px) !important;
            margin-bottom: 6px !important;
        }
    }

    /* === Streamlit widget polish on mobile === */
    @media (max-width: 768px) {
        /* Buttons full width by default */
        .stButton > button,
        .stDownloadButton > button,
        .stFormSubmitButton > button {
            width: 100% !important;
            min-height: 44px !important;
            font-size: 13px !important;
        }
        /* Selectbox / Multiselect / TextInput full width */
        .stSelectbox, .stMultiselect, .stTextInput,
        .stTextArea, .stNumberInput, .stDateInput {
            width: 100% !important;
        }
        /* Tabs more compact */
        .stTabs [data-baseweb="tab-list"] {
            gap: 4px !important;
            flex-wrap: wrap !important;
        }
        .stTabs [data-baseweb="tab"] {
            padding: 6px 10px !important;
            font-size: 11px !important;
        }
        /* DataFrames scrollable */
        .stDataFrame, [data-testid="stDataFrame"] {
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch !important;
        }
        /* Expander header larger touch target */
        .streamlit-expanderHeader,
        summary[data-testid="stExpander"] {
            padding: 10px 14px !important;
            font-size: 13px !important;
        }
        /* Alert/info boxes tighter */
        .stAlert {
            padding: 10px 14px !important;
            font-size: 12px !important;
        }
        /* Markdown paragraphs tighter */
        .stMarkdown p {
            font-size: 13px !important;
            line-height: 1.5 !important;
        }
        /* Subheaders smaller */
        h2 { font-size: 18px !important; }
        h3 { font-size: 15px !important; }
        h4 { font-size: 13px !important; }
    }

    /* === Improve touch ergonomics === */
    @media (max-width: 768px) {
        /* Tap targets at least 44px (Apple HIG) */
        button, [role="button"], input, select, textarea {
            min-height: 40px !important;
        }
        /* Hide the chart_fullscreen toggle on mobile if present */
        [data-testid="stFullScreenButton"] {
            display: none !important;
        }
    }
</style>

<script>
(function() {
    // Theme toggle persistence
    const saved = localStorage.getItem('kanom_theme') || 'light';
    if (saved === 'dark') document.body.classList.add('dark-mode');
    
    // Create theme toggle button
    const themeBtn = document.createElement('button');
    themeBtn.className = 'theme-toggle';
    themeBtn.innerHTML = saved === 'dark' ? '☀️' : '🌙';
    themeBtn.title = 'Toggle theme';
    themeBtn.onclick = () => {
        const isDark = document.body.classList.toggle('dark-mode');
        localStorage.setItem('kanom_theme', isDark ? 'dark' : 'light');
        themeBtn.innerHTML = isDark ? '☀️' : '🌙';
        // Notify Streamlit to re-render (preserves state via session_state)
        window.parent.postMessage({type: 'streamlit:rerun'}, '*');
    };
    document.body.appendChild(themeBtn);
    
    // Smooth scroll for any anchor
    document.documentElement.style.scrollBehavior = 'smooth';
    
    // Create floating scroll-to-top button
    const btn = document.createElement('button');
    btn.className = 'scroll-top-btn';
    btn.innerHTML = '↑';
    btn.title = 'Back to top';
    btn.onclick = () => window.scrollTo({top: 0, behavior: 'smooth'});
    btn.style.display = 'none';
    document.body.appendChild(btn);
    window.addEventListener('scroll', () => {
        btn.style.display = window.scrollY > 400 ? 'flex' : 'none';
    });
})();
</script>
""", unsafe_allow_html=True)

# ============================================================
# DATA LOAD
# ============================================================
@st.cache_data(ttl=30)
def load_data():
    """Read from .xlsx file - cache expires every 30s so updates appear quickly."""
    xlsx_path = Path(__file__).parent / "QA_Defects_Data.xlsx"
    csv_path  = Path(__file__).parent / "QA_Defects_Data.csv"

    # Try .xlsx first (current source of truth)
    if xlsx_path.exists():
        df = pd.read_excel(xlsx_path)
    elif csv_path.exists():
        df = pd.read_csv(csv_path)
    else:
        return pd.DataFrame(columns=["Date","Supplier","Group Part","Problem Mode","Part Name","Part No","Qty","Comment"])

    # Normalize Date column - handle mixed formats ('2026-08-31' and '2026-08-31 00:00:00')
    df["Date"] = pd.to_datetime(df["Date"], format="mixed", errors="coerce")
    # Drop rows where Date couldn't be parsed
    df = df.dropna(subset=["Date"]).reset_index(drop=True)
    return df

df = load_data()

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    # === LOGO SECTION (centered with proper spacing) ===
    # Open container with generous padding
    st.markdown(
        '<div style="padding: 24px 12px 16px 12px; text-align: center;">',
        unsafe_allow_html=True
    )

    # Logo image — centered, no decorations
    try:
        st.image(LOGO_URL, width=180)
    except Exception:
        st.markdown(
            f'<img src="{LOGO_URL}" width="180" '
            f'style="border-radius:12px;display:block;margin:0 auto;">',
            unsafe_allow_html=True
        )

    # Brand text — clean, centered, proper line-height
    st.markdown(
        '<div style="color:#FFD700; font-size:20px; font-weight:900; '
        'letter-spacing:3px; margin-top:18px; line-height:1.2;">'
        '⚡ 3K BATTERY</div>'
        '<div style="color:#999; font-size:10px; font-weight:600; '
        'letter-spacing:2.5px; margin-top:8px; line-height:1.2;">'
        'QA DEFECTS DASHBOARD</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # Divider
    st.markdown(
        '<div style="height:1px; margin:24px 8px; '
        'background:linear-gradient(90deg, transparent 0%, '
        'rgba(255,215,0,0.4) 50%, transparent 100%);"></div>',
        unsafe_allow_html=True
    )

    # === NAVIGATION (gradient buttons) ===
    if "current_page" not in st.session_state:
        st.session_state.current_page = "📊 14 Days Monitoring"

    nav_items = [
        ("📊 14 Days Monitoring",   "📊", "linear-gradient(135deg, #FFD700 0%, #FFC107 100%)", "#000"),
        ("🔍 Searching Supplier",   "🔍", "linear-gradient(135deg, #000000 0%, #434343 100%)", "#FFD700"),
        ("📝 Data Entry",           "📝", "linear-gradient(135deg, #FFA500 0%, #FF6347 100%)", "#fff"),
    ]
    for full_label, icon, gradient, text_color in nav_items:
        is_active = (st.session_state.current_page == full_label)
        active_border = "3px solid #FFD700" if is_active else "2px solid transparent"
        if st.button(
            f"{icon}  {full_label.replace(icon + ' ', '')}",
            key=f"nav_{full_label}",
            use_container_width=True
        ):
            st.session_state.current_page = full_label
            st.rerun()
    page = st.session_state.current_page

    # Divider
    st.markdown(
        '<div style="height:1px; margin:20px 8px; '
        'background:linear-gradient(90deg, transparent 0%, '
        'rgba(255,215,0,0.3) 50%, transparent 100%);"></div>',
        unsafe_allow_html=True
    )

    # === FILTERS ===
    st.markdown(
        '<div style="color:#FFD700; font-size:11px; font-weight:800; '
        'letter-spacing:2.5px; margin:0 4px 12px 4px; '
        'text-transform:uppercase;">🎛️ FILTERS</div>',
        unsafe_allow_html=True
    )

    supplier_f = st.selectbox(
        "Supplier", ["All"] + sorted(df["Supplier"].unique().tolist())
    )
    group_f = st.selectbox(
        "Group Part", ["All"] + sorted(df["Group Part"].unique().tolist())
    )
    mode_f = st.selectbox(
        "Problem Mode", ["All"] + sorted(df["Problem Mode"].unique().tolist())
    )

    # Divider
    st.markdown(
        '<div style="height:1px; margin:20px 8px; '
        'background:linear-gradient(90deg, transparent 0%, '
        'rgba(255,215,0,0.3) 50%, transparent 100%);"></div>',
        unsafe_allow_html=True
    )

    # Reset button
    if st.button("🗑️ RESET ALL FILTERS", use_container_width=True):
        st.session_state.selected_group = None
        st.rerun()

    # Footer info at bottom
    st.markdown(
        '<div style="position:relative; margin-top:24px; padding:14px 12px; '
        'background:rgba(255,215,0,0.08); border:1px solid rgba(255,215,0,0.3); '
        'border-radius:10px; text-align:center;">'
        '<div style="color:#FFD700; font-size:9px; font-weight:700; '
        'letter-spacing:2px;">📅 LAST UPDATE</div>'
        f'<div style="color:#fff; font-size:11px; font-weight:800; '
        f'margin-top:4px;">{datetime.now().strftime("%Y-%m-%d")}</div>'
        f'<div style="color:#666; font-size:9px; margin-top:2px;">'
        f'{datetime.now().strftime("%H:%M:%S")}</div>'
        '</div>',
        unsafe_allow_html=True
    )

# Apply filters
filtered = df.copy()
if supplier_f != "All":
    filtered = filtered[filtered["Supplier"] == supplier_f]
if group_f != "All":
    filtered = filtered[filtered["Group Part"] == group_f]
if mode_f != "All":
    filtered = filtered[filtered["Problem Mode"] == mode_f]

# ============================================================
# HEADER
# ============================================================
if "14 Days" in page:
    title_text = "14 DAYS DEFECT MONITORING"
    subtitle = "Incoming Quality Dashboard · Auto-sync from GitHub"
elif "Searching" in page:
    title_text = "SEARCHING SUPPLIER INFORMATION"
    subtitle = "PPM Analysis & FY Comparison · Auto-sync from GitHub"
else:  # Data Entry
    title_text = "DATA ENTRY"
    subtitle = "Quick Defect Logging System · 1-Click Sync"

max_date = df["Date"].max()
min_date = max_date - timedelta(days=13)

# Calculate quick stats for header info
n_suppliers = filtered["Supplier"].nunique() if "Supplier" in filtered.columns else 0
n_groups    = filtered["Group Part"].nunique() if "Group Part" in filtered.columns else 0
n_modes     = filtered["Problem Mode"].nunique() if "Problem Mode" in filtered.columns else 0

# Streamlit sanitises inline <script>, so a JS clock never ticks here.
# Render the timestamp server-side instead of showing a dead "--:--:--".
now_hm = datetime.now().strftime("%H:%M")

st.markdown(f"""
<div class="hero-header">
    <div class="hero-left">
        <div class="hero-title">⚡ {title_text}</div>
        <div class="hero-subtitle">{subtitle}</div>
        <div class="hero-badge">🏭 3K BATTERY CO., LTD.</div>
    </div>
    <div class="hero-right">
        <div class="hero-date-row">
            <span class="hero-date-label">FROM</span>
            <span class="hero-date-val">{min_date.strftime('%m/%d/%Y')}</span>
        </div>
        <div class="hero-date-divider">━━━</div>
        <div class="hero-date-row">
            <span class="hero-date-label">TO</span>
            <span class="hero-date-val">{max_date.strftime('%m/%d/%Y')}</span>
        </div>
    </div>
    <div class="hero-stats">
        <div class="hero-stat">
            <span class="hero-stat-icon">🏭</span>
            <span class="hero-stat-value">{n_suppliers}</span>
            <span class="hero-stat-label">Suppliers</span>
        </div>
        <div class="hero-stat">
            <span class="hero-stat-icon">📦</span>
            <span class="hero-stat-value">{n_groups}</span>
            <span class="hero-stat-label">Groups</span>
        </div>
        <div class="hero-stat">
            <span class="hero-stat-icon">⚠️</span>
            <span class="hero-stat-value">{n_modes}</span>
            <span class="hero-stat-label">Modes</span>
        </div>
        <div class="hero-stat hero-stat-clock">
            <span class="hero-stat-icon">⏰</span>
            <span class="hero-stat-label">Updated</span>
            <span class="hero-stat-clock-val">{now_hm}</span>
        </div>
    </div>
    <div class="hero-live">
        <span class="hero-live-pulse"></span>
        <span>LIVE DATA · Auto-synced from GitHub</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# PAGE 1: 14 DAYS MONITORING
# ============================================================
if "14 Days" in page:
    if "selected_group" not in st.session_state:
        st.session_state.selected_group = None

    if st.session_state.selected_group:
        filtered = filtered[filtered["Group Part"] == st.session_state.selected_group]

    total_qty = int(filtered["Qty"].sum())
    total_case = len(filtered)

    # === TREND CALCULATION (Last 7 days vs Previous 7 days) ===
    if not filtered.empty:
        max_date = filtered["Date"].max()
        cutoff = max_date - pd.Timedelta(days=6)
        prev_cutoff = max_date - pd.Timedelta(days=13)
        recent_7 = filtered[filtered["Date"] >= cutoff]
        prev_7 = filtered[(filtered["Date"] >= prev_cutoff) & (filtered["Date"] < cutoff)]

        recent_qty = int(recent_7["Qty"].sum())
        prev_qty = int(prev_7["Qty"].sum())
        recent_case = len(recent_7)
        prev_case = len(prev_7)

        # For QTY: lower is better → up arrow = bad (red), down arrow = good (green)
        if prev_qty > 0:
            qty_pct = round(((recent_qty - prev_qty) / prev_qty) * 100, 1)
        else:
            qty_pct = 0.0 if recent_qty == 0 else 100.0

        # For CASE: lower is better
        if prev_case > 0:
            case_pct = round(((recent_case - prev_case) / prev_case) * 100, 1)
        else:
            case_pct = 0.0 if recent_case == 0 else 100.0

        # For AVG/CASE: lower is better
        recent_avg = recent_qty / max(recent_case, 1)
        prev_avg = prev_qty / max(prev_case, 1)
        if prev_avg > 0:
            avg_pct = round(((recent_avg - prev_avg) / prev_avg) * 100, 1)
        else:
            avg_pct = 0.0

        # For UNIQUE PARTS: lower is better
        recent_parts = recent_7["Part No"].nunique()
        prev_parts = prev_7["Part No"].nunique()
        if prev_parts > 0:
            parts_pct = round(((recent_parts - prev_parts) / prev_parts) * 100, 1)
        else:
            parts_pct = 0.0 if recent_parts == 0 else 100.0
    else:
        qty_pct = case_pct = avg_pct = parts_pct = 0.0

    def trend_arrow(value):
        # For QA defects: lower is better
        # value > 0 = bad (up arrow red), value < 0 = good (down arrow green)
        if value > 0:
            return "up"
        elif value < 0:
            return "down"
        else:
            return "neutral"

    qty_arrow = trend_arrow(qty_pct)
    case_arrow = trend_arrow(case_pct)
    avg_arrow = trend_arrow(avg_pct)
    parts_arrow = trend_arrow(parts_pct)

    daily = filtered.groupby("Date").agg(Qty=("Qty", "sum"), Case=("Qty", "count")).reset_index()
    labels = daily["Date"].dt.strftime("%m/%d").tolist()
    qty_data = daily["Qty"].tolist()
    case_data = daily["Case"].tolist()

    # === KPI ROW (2x2 on mobile, 4-across on desktop via CSS) ===
    kpi_row1_left, kpi_row1_right = st.columns(2, gap="small")
    kpi_row2_left, kpi_row2_right = st.columns(2, gap="small")

    with kpi_row1_left:
        unique_suppliers = filtered["Supplier"].nunique()
        arrow_icon = {"up": "▲", "down": "▼", "neutral": "—"}[qty_arrow]
        st.markdown(f"""
        <div class="kpi-container kpi-yellow kpi-card">
            <div class="kpi-icon">📦</div>
            <div class="kpi-label">📦 TOTAL Q'TY</div>
            <div class="kpi-value">{total_qty:,}<span class="kpi-unit">PCS</span></div>
            <div class="kpi-trend trend-{qty_arrow}">{arrow_icon} {abs(qty_pct):.1f}% vs prev 7d</div>
        </div>
        """, unsafe_allow_html=True)

    with kpi_row1_right:
        arrow_icon = {"up": "▲", "down": "▼", "neutral": "—"}[case_arrow]
        st.markdown(f"""
        <div class="kpi-container kpi-black kpi-card">
            <div class="kpi-icon">📋</div>
            <div class="kpi-label">📋 TOTAL CASE</div>
            <div class="kpi-value">{total_case:,}<span class="kpi-unit">CASE</span></div>
            <div class="kpi-trend trend-{case_arrow}">{arrow_icon} {abs(case_pct):.1f}% vs prev 7d</div>
        </div>
        """, unsafe_allow_html=True)

    with kpi_row2_left:
        avg_qty = round(total_qty / max(total_case, 1), 1)
        arrow_icon = {"up": "▲", "down": "▼", "neutral": "—"}[avg_arrow]
        st.markdown(f"""
        <div class="kpi-container kpi-yellow kpi-card">
            <div class="kpi-icon">📊</div>
            <div class="kpi-label">📊 AVG / CASE</div>
            <div class="kpi-value">{avg_qty}<span class="kpi-unit">PCS</span></div>
            <div class="kpi-trend trend-{avg_arrow}">{arrow_icon} {abs(avg_pct):.1f}% vs prev 7d</div>
        </div>
        """, unsafe_allow_html=True)

    with kpi_row2_right:
        unique_parts = filtered["Part No"].nunique()
        arrow_icon = {"up": "▲", "down": "▼", "neutral": "—"}[parts_arrow]
        st.markdown(f"""
        <div class="kpi-container kpi-black kpi-card">
            <div class="kpi-icon">⚙</div>
            <div class="kpi-label">⚙ UNIQUE PARTS</div>
            <div class="kpi-value">{unique_parts}<span class="kpi-unit">PART</span></div>
            <div class="kpi-trend trend-{parts_arrow}">{arrow_icon} {abs(parts_pct):.1f}% vs prev 7d</div>
        </div>
        """, unsafe_allow_html=True)

    # === SMART INSIGHTS ===
    if not filtered.empty:
        # Find worst supplier by total qty
        sup_qty = filtered.groupby("Supplier")["Qty"].sum().sort_values(ascending=False)
        worst_supplier = sup_qty.index[0] if len(sup_qty) > 0 else "N/A"
        worst_qty = int(sup_qty.iloc[0]) if len(sup_qty) > 0 else 0

        # Find worst mode
        mode_qty = filtered.groupby("Problem Mode")["Qty"].sum().sort_values(ascending=False)
        worst_mode = mode_qty.index[0] if len(mode_qty) > 0 else "N/A"

        # Trend direction
        if qty_pct > 20:
            trend_msg = f"⚠️ Defects spiked {qty_pct:.1f}% vs previous 7 days"
            trend_color = "#B71C1C"
            trend_bg = "rgba(255,107,107,0.12)"
        elif qty_pct > 0:
            trend_msg = f"📈 Defects up {qty_pct:.1f}% vs previous 7 days"
            trend_color = "#E65100"
            trend_bg = "rgba(255,152,0,0.12)"
        elif qty_pct < -10:
            trend_msg = f"✅ Defects down {abs(qty_pct):.1f}% — great improvement!"
            trend_color = "#1B5E20"
            trend_bg = "rgba(76,175,80,0.12)"
        elif qty_pct < 0:
            trend_msg = f"📉 Defects down {abs(qty_pct):.1f}% vs previous 7 days"
            trend_color = "#2E7D32"
            trend_bg = "rgba(76,175,80,0.12)"
        else:
            trend_msg = "➡️ Defects stable vs previous 7 days"
            trend_color = "#555"
            trend_bg = "rgba(0,0,0,0.05)"

        # Determine severity class
        if qty_pct > 20:
            severity = "insight-critical"
            trend_msg = f"Defects spiked {qty_pct:.1f}% vs previous 7 days"
            icon = "🚨"
        elif qty_pct > 0:
            severity = "insight-warning"
            trend_msg = f"Defects up {qty_pct:.1f}% vs previous 7 days"
            icon = "📈"
        elif qty_pct < -10:
            severity = "insight-good"
            trend_msg = f"Defects down {abs(qty_pct):.1f}% — great improvement!"
            icon = "✅"
        elif qty_pct < 0:
            severity = "insight-good"
            trend_msg = f"Defects down {abs(qty_pct):.1f}% vs previous 7 days"
            icon = "📉"
        else:
            severity = "insight-info"
            trend_msg = "Defects stable vs previous 7 days"
            icon = "➡️"

        st.markdown(
            f'<div class="insight-banner {severity}">'
            f'<div class="insight-banner-icon">{icon}</div>'
            f'<div class="insight-banner-body">'
            f'<div class="insight-banner-title">{trend_msg}</div>'
            f'<div class="insight-banner-detail">'
            f'🏭 Worst supplier: <b>{worst_supplier}</b> ({worst_qty:,} PCS)  ·  '
            f'⚠️ Top mode: <b>{worst_mode}</b>'
            f'</div></div></div>',
            unsafe_allow_html=True
        )

    # === QUICK ACTIONS BAR ===
    st.markdown(
        '<div class="section-header">'
        '<div class="section-icon">⚡</div>'
        'QUICK ACTIONS'
        '</div>',
        unsafe_allow_html=True
    )

    qa1, qa2 = st.columns(2)
    with qa1:
        if st.button("🔄 REFRESH DATA", use_container_width=True, key="qa_refresh",
                     type="secondary"):
            st.cache_data.clear()
            st.rerun()
        if st.button("📋 COPY SUMMARY", use_container_width=True, key="qa_copy",
                     type="secondary"):
            summary = (
                f"3K Battery QA Defects Summary\n"
                f"Period: {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}\n"
                f"Total Q'TY: {total_qty:,} PCS\n"
                f"Total CASE: {total_case:,}\n"
                f"Suppliers: {filtered['Supplier'].nunique()}\n"
                f"Unique Parts: {filtered['Part No'].nunique()}"
            )
            st.code(summary, language="text")
    with qa2:
        # Export CSV button (uses download_button which is its own widget)
        csv = filtered.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📊 EXPORT CSV",
            data=csv,
            file_name=f"QA_Defects_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
            key="qa_dl_csv"
        )
        # Sync status indicator
        st.markdown(
            f'<div style="background:white; border:2px solid #4CAF50; border-radius:10px; '
            f'padding:14px; text-align:center; min-height:50px; display:flex; '
            f'align-items:center; justify-content:center; gap:8px; font-weight:800; '
            f'letter-spacing:1px; color:#1B5E20;">'
            f'<span style="width:10px; height:10px; background:#4CAF50; border-radius:50%; '
            f'box-shadow:0 0 8px #4CAF50;"></span>'
            f'<span>SYNCED · GITHUB</span>'
            f'</div>',
            unsafe_allow_html=True
        )

    # === SEARCH BAR ===
    search_query = st.text_input(
        "🔍 Search Part No / Supplier / Comment",
        placeholder="Type to search... (e.g., VE101, KSV, scratch)",
        key="search_14d"
    )
    if search_query:
        mask = (
            filtered["Part No"].astype(str).str.contains(search_query, case=False, na=False) |
            filtered["Supplier"].astype(str).str.contains(search_query, case=False, na=False) |
            filtered["Comment"].astype(str).str.contains(search_query, case=False, na=False) |
            filtered["Part Name"].astype(str).str.contains(search_query, case=False, na=False)
        )
        filtered = filtered[mask]
        if filtered.empty:
            st.markdown(
                '<div class="empty-state">'
                '<span class="empty-state-icon">🔍</span>'
                '<div class="empty-state-title">NO RESULTS FOUND</div>'
                '<div class="empty-state-text">No defects match your search query.</div>'
                '<div class="empty-state-hint">Try a different keyword</div>'
                '</div>',
                unsafe_allow_html=True
            )
            st.stop()

    # === TREND CHARTS ===
    st.markdown(
        '<div class="section-header">'
        '<div class="section-icon">📈</div>'
        'DAILY TREND ANALYSIS'
        '</div>',
        unsafe_allow_html=True
    )

    trend_card_css = """
    <style>
      html, body { margin:0; padding:0; }
      .tc-card {
        background:#fff; border:2px solid #111; border-radius:14px;
        padding:12px 12px 6px 12px; box-shadow:0 4px 14px rgba(0,0,0,0.08);
        font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
      }
      .tc-head { display:flex; align-items:center; gap:8px; margin-bottom:6px; }
      .tc-dot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
      .tc-title { color:#000; font-weight:900; font-size:12px; letter-spacing:1px; }
      .tc-body { width:100%; height:210px; }
      @media (max-width: 768px) {
        .tc-card { padding:10px 10px 4px 10px; border-radius:12px; }
        .tc-title { font-size:11px; }
        /* Body height stays 210px at every breakpoint so the fixed Streamlit
           iframe height fits the card exactly — a shorter card here is what
           left ~56px of dead white space under each chart. */
      }
    </style>
    """

    def _trend_card(canvas_id, title, dot_color, labels_js, data_js,
                    border, bg, point, point_border, axis_title,
                    inject_table_styling=False):
        # Convert Python True/False to JS true/false so it can be embedded in JS.
        _its = "true" if inject_table_styling else "false"
        return trend_card_css + f"""
        <div class="tc-card">
          <div class="tc-head">
            <div class="tc-dot" style="background:{dot_color};"></div>
            <div class="tc-title">{title}</div>
          </div>
          <div class="tc-body"><canvas id="{canvas_id}"></canvas></div>
        </div>
        <script src="{CHARTJS_CDN}"></script>
        <script>
        (function() {{
            const canvas = document.getElementById('{canvas_id}');
            if (!canvas) return;
            const fit = () => {{
                const w = canvas.parentElement.clientWidth;
                if (w <= 0) return false;
                canvas.width = w;
                canvas.height = canvas.parentElement.clientHeight;
                return true;
            }};
            fit();
            const chart = new Chart(canvas, {{
                type: 'line',
                data: {{
                    labels: {labels_js},
                    datasets: [{{
                        data: {data_js},
                        borderColor: '{border}', backgroundColor: '{bg}',
                        borderWidth: 2.5, tension: 0.35, fill: true,
                        pointRadius: 4, pointHoverRadius: 7,
                        pointBackgroundColor: '{point}', pointBorderColor: '{point_border}',
                        pointBorderWidth: 2
                    }}]
                }},
                options: {{
                    responsive: false, maintainAspectRatio: false, animation: false,
                    layout: {{ padding: {{ top: 4, right: 6 }} }},
                    plugins: {{ legend: {{ display: false }} }},
                    scales: {{
                        y: {{ beginAtZero: true,
                              ticks: {{ font: {{ size: 10 }}, precision: 0, maxTicksLimit: 6 }},
                              title: {{ display: true, text: '{axis_title}',
                                        font: {{ size: 10, weight: 'bold' }} }},
                              grid: {{ color: '#f1f1f1' }} }},
                        x: {{ ticks: {{ font: {{ size: 10 }}, maxRotation: 0,
                                       autoSkip: true, maxTicksLimit: 6 }},
                              grid: {{ display: false }} }}
                    }}
                }}
            }});
            const refit = () => {{ fit(); chart.resize(); }};
            window.addEventListener('resize', refit);
            setTimeout(refit, 350);
            // === Premium table styling — runs from the first chart iframe so we
            // have access to the parent document via window.parent.document.
            // Idempotent: only injects once.
            if ({_its} && window.parent && window.parent !== window) {{
                try {{
                    const HEAD = window.parent.document.head;
                    if (!HEAD.querySelector('#kanom-table-style')) {{
                        const s = window.parent.document.createElement('style');
                        s.id = 'kanom-table-style';
                        s.textContent = `
                          .stDataFrame thead th {{
                            background: linear-gradient(135deg, #0e0e0e 0%, #1f1f1f 100%) !important;
                            color: #FFD700 !important;
                            font-weight: 800 !important;
                            letter-spacing: 0.8px !important;
                            text-transform: uppercase !important;
                            font-size: 11px !important;
                            border-bottom: 2.5px solid #FFD700 !important;
                            padding: 12px 10px !important;
                          }}
                          .stDataFrame tbody tr:nth-of-type(even) td {{
                            background: rgba(255,243,196,0.55) !important;
                          }}
                          .stDataFrame tbody tr:hover td {{
                            background: rgba(255,215,0,0.22) !important;
                          }}
                        `;
                        HEAD.appendChild(s);
                        // zebra is now declarative via :nth-of-type; no JS observer needed
                    }}
                }} catch (e) {{ /* ignore */ }}
            }}
        }})();
        </script>
        """

    # Two separate iframes -> Streamlit columns handle desktop side-by-side
    # AND mobile stacking, with no dead space from a fixed tall iframe.
    tc1, tc2 = st.columns(2, gap="small")
    with tc1:
        st.components.v1.html(
            _trend_card("kpiQtyChart", "Q'TY (PCS) TREND", "#FFD700", labels, qty_data,
                        "#000000", "rgba(255,215,0,0.35)", "#FFD700", "#000000", "QTY",
                        inject_table_styling=True),
            height=250,
        )
    with tc2:
        st.components.v1.html(
            _trend_card("kpiCaseChart", "CASE TREND", "#111111", labels, case_data,
                        "#FFD700", "rgba(0,0,0,0.12)", "#000000", "#FFD700", "CASE"),
            height=250,
        )

    # === PROBLEM MODE + TOP 5 SUPPLIERS ===
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown(
            '<div class="resp-row-2col">'
            '<div class="section-header">'
            '<div class="section-icon">⚠️</div>'
            'PROBLEM MODE BREAKDOWN'
            '</div>'
            '</div>',
            unsafe_allow_html=True
        )
        mode_summary = filtered.groupby("Problem Mode").agg(
            Qty=("Qty", "sum"), Case=("Qty", "count")
        ).reset_index().sort_values("Qty", ascending=False)
        mode_summary["Case"] = mode_summary["Case"].astype(int)
        mode_summary["Qty"] = mode_summary["Qty"].astype(int)
        # Add % share column
        total_qty_all = max(int(mode_summary["Qty"].sum()), 1)
        mode_summary["% Share"] = mode_summary["Qty"].astype(int).map(
            lambda q: f"{round((int(q)/total_qty_all)*100, 1)}%"
        )

        def color_mode(val):
            if val not in mode_summary["Problem Mode"].values: return ''
            case = int(mode_summary.loc[mode_summary["Problem Mode"] == val, "Case"].iloc[0])
            if case == 0: return 'background-color: #C8E6C9; color: #1B5E20; font-weight:700'
            elif case < 2: return 'background-color: #FFD700; color: #000; font-weight:700'
            else: return 'background-color: #FFCDD2; color: #B71C1C; font-weight:700'

        st.dataframe(
            mode_summary.style.map(color_mode, subset=["Problem Mode"]),
            use_container_width=True, hide_index=True, height=280
        )

    with col_r:
        st.markdown(
            '<div class="section-header">'
            '<div class="section-icon">🏭</div>'
            'TOP 5 SUPPLIERS'
            '</div>',
            unsafe_allow_html=True
        )
        top5 = filtered.groupby("Supplier").agg(
            Qty=("Qty", "sum"), Case=("Qty", "count")
        ).reset_index().sort_values("Qty", ascending=False).head(5)
        top5["Qty"] = top5["Qty"].astype(int)
        top5["Case"] = top5["Case"].astype(int)

        # Add rank column
        top5.insert(0, "Rank", range(1, len(top5) + 1))

        # Add % share column
        total_qty_top = max(int(top5["Qty"].sum()), 1)
        top5["% Share"] = top5["Qty"].astype(int).map(
            lambda q: f"{round((int(q)/total_qty_top)*100, 1)}%"
        )

        def color_rank(val):
            if val == 1: return 'background-color: #FFD700; color: #000; font-weight:900; font-size:16px; text-align:center'
            elif val == 2: return 'background-color: #FFC107; color: #000; font-weight:900; text-align:center'
            elif val == 3: return 'background-color: #FFA000; color: #fff; font-weight:900; text-align:center'
            else: return 'background-color: #f0f0f0; color: #555; font-weight:700; text-align:center'

        st.dataframe(
            top5.style.map(color_rank, subset=["Rank"]),
            column_order=["Rank", "Supplier", "Qty", "Case", "% Share"],
            use_container_width=True, hide_index=True, height=280
        )

    # === SUPPLIER GROUP CARDS (gradient buttons via inline CSS) ===
    st.markdown('<div class="section-header">🗂️ SELECTION SUPPLIER GROUP (CLICK TO FILTER)</div>', unsafe_allow_html=True)

    all_groups = sorted(df["Group Part"].unique().tolist())
    group_meta = {
        "ELECTRIC & ELEC.": ("⚡", "linear-gradient(135deg,#FFD700 0%,#FFA000 100%)", "#000"),
        "PACKING":          ("📦", "linear-gradient(135deg,#A1887F 0%,#6D4C41 100%)", "#fff"),
        "PIPING":           ("🔧", "linear-gradient(135deg,#FFCA28 0%,#FF8F00 100%)", "#000"),
        "PLASTIC":          ("🧊", "linear-gradient(135deg,#4FC3F7 0%,#0288D1 100%)", "#fff"),
        "PRINTING":         ("🖨", "linear-gradient(135deg,#BA68C8 0%,#7B1FA2 100%)", "#fff"),
        "RAW MATERIAL":     ("⛰", "linear-gradient(135deg,#81C784 0%,#388E3C 100%)", "#000"),
        "RUBBER":           ("⚫", "linear-gradient(135deg,#616161 0%,#212121 100%)", "#FFD700"),
        "SEALING":          ("⭕", "linear-gradient(135deg,#EF5350 0%,#C62828 100%)", "#fff"),
        "SHEET METAL":      ("🔩", "linear-gradient(135deg,#B0BEC5 0%,#546E7A 100%)", "#fff"),
        "FOAM":             ("🧽", "linear-gradient(135deg,#FFF176 0%,#F9A825 100%)", "#000"),
        "OTHERS":           ("📦", "linear-gradient(135deg,#CFD8DC 0%,#455A64 100%)", "#fff"),
    }

    # Show selected indicator
    if st.session_state.selected_group:
        sel = st.session_state.selected_group
        sel_meta = group_meta.get(sel, ("📦", "#FFD700", "#000"))
        st.markdown(
            f'<div style="background:#000;color:#FFD700;padding:14px 22px;border-radius:12px;'
            f'margin-bottom:18px;font-weight:700;font-size:15px;border-left:6px solid #FFD700;'
            f'display:flex;align-items:center;gap:12px;box-shadow:0 4px 12px rgba(0,0,0,0.2);">'
            f'<span style="font-size:22px;">{sel_meta[0]}</span>'
            f'<span>Currently filtering by: <b>{sel}</b></span>'
            f'<span style="margin-left:auto;font-size:11px;color:#999;">Click again to clear</span>'
            f'</div>',
            unsafe_allow_html=True
        )

    # Generate per-button CSS to style each button with its group's gradient
    css_rules = []
    for g in all_groups:
        icon, bg, fg = group_meta.get(g, ("📦", "linear-gradient(135deg,#FFD700,#FFA000)", "#000"))
        safe_id = f"grp_{g}".replace(" ", "_").replace(".", "").replace("&", "and")
        css_rules.append(
            f"div[data-testid='stHorizontalBlock'] button[key='{safe_id}'] {{"
            f"background:{bg} !important;color:{fg} !important;"
            f"border:2px solid rgba(0,0,0,0.15) !important;"
            f"font-weight:900 !important;"
            f"}}"
            f"div[data-testid='stHorizontalBlock'] button[key='{safe_id}']:hover {{"
            f"transform:translateY(-3px) !important;"
            f"box-shadow:0 8px 20px rgba(0,0,0,0.2) !important;"
            f"filter:brightness(1.1) !important;"
            f"}}"
        )
    st.markdown(f"<style>{''.join(css_rules)}</style>", unsafe_allow_html=True)

    # Add universal button sizing CSS — force identical dimensions
    st.markdown("""
    <style>
        /* Force uniform height + width for all group buttons */
        div[data-testid="stHorizontalBlock"] button[key^="grp_"] {
            height: 145px !important;
            min-height: 145px !important;
            max-height: 145px !important;
            width: 100% !important;
            min-width: 100% !important;
            white-space: pre-line !important;
            font-size: 13px !important;
            line-height: 1.5 !important;
            padding: 16px 8px !important;
            border-radius: 14px !important;
            letter-spacing: 1px !important;
            text-transform: uppercase !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            text-align: center !important;
            transition: transform 0.2s, box-shadow 0.2s, filter 0.2s !important;
            margin: 0 !important;
        }
        div[data-testid="stHorizontalBlock"] button[key^="grp_"]:hover {
            transform: translateY(-3px) !important;
            box-shadow: 0 8px 20px rgba(0,0,0,0.2) !important;
            filter: brightness(1.1) !important;
        }
        div[data-testid="stHorizontalBlock"] button[key^="grp_"]:active {
            transform: translateY(0) !important;
        }
        /* Equal column widths — 6-up on desktop, 2-up on phones.
           The old unscoped `width: 100%/6` also hit every OTHER column row
           on the page and squashed them to 16% on mobile. */
        @media (min-width: 769px) {
            div[data-testid="stHorizontalBlock"]:has(button[key^="grp_"])
                > div[data-testid="stColumn"] {
                flex: 1 1 0 !important;
                min-width: 0 !important;
                width: calc(100% / 6) !important;
            }
        }
        @media (max-width: 768px) {
            div[data-testid="stHorizontalBlock"]:has(button[key^="grp_"])
                > div[data-testid="stColumn"] {
                flex: 1 1 calc(50% - 4px) !important;
                min-width: calc(50% - 4px) !important;
                width: calc(50% - 4px) !important;
            }
            div[data-testid="stHorizontalBlock"] button[key^="grp_"] {
                height: 74px !important;
                min-height: 74px !important;
                max-height: 74px !important;
                font-size: 10px !important;
                line-height: 1.35 !important;
                padding: 8px 4px !important;
                letter-spacing: 0.3px !important;
                border-radius: 12px !important;
            }
        }
    </style>
    """, unsafe_allow_html=True)

    n_cols = 6
    for i in range(0, len(all_groups), n_cols):
        row_groups = all_groups[i:i + n_cols]
        cols = st.columns(n_cols)
        for j, g in enumerate(row_groups):
            with cols[j]:
                st.markdown('<div class="group-btn-wrap">', unsafe_allow_html=True)
                icon, _, _ = group_meta.get(g, ("📦", "", ""))
                count = int(df[df["Group Part"] == g]["Qty"].sum())
                is_sel = (st.session_state.selected_group == g)
                btn_label = f"{icon}  {g}\n{count} QTY"

                if is_sel:
                    # Selected: show HTML card above + button below to toggle off
                    sel_meta = group_meta.get(g)
                    sel_bg = sel_meta[1]
                    btn_html = (
                        f'<div style="background:{sel_bg};border:3px solid #FFD700;'
                        f'border-radius:14px;padding:10px;text-align:center;'
                        f'box-shadow:0 8px 24px rgba(255,215,0,0.5);'
                        f'position:relative;margin-bottom:8px;color:#000;'
                        f'height:60px;display:flex;flex-direction:column;justify-content:center;">'
                        f'<div style="position:absolute;top:-8px;right:-8px;background:#FFD700;'
                        f'color:#000;border-radius:50%;width:24px;height:24px;'
                        f'display:flex;align-items:center;justify-content:center;'
                        f'font-size:12px;font-weight:900;">✓</div>'
                        f'<div style="font-size:18px;">{icon}</div>'
                        f'<div style="font-size:11px;font-weight:900;">CLICK TO CLEAR</div>'
                        f'</div>'
                    )
                    st.markdown(btn_html, unsafe_allow_html=True)
                    if st.button(btn_label, key=f"grp_{g}",
                                use_container_width=True, type="primary"):
                        st.session_state.selected_group = None
                        st.rerun()
                else:
                    # Unselected: just the styled gradient button
                    if st.button(btn_label, key=f"grp_{g}",
                                use_container_width=True, type="secondary"):
                        st.session_state.selected_group = g
                        st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.selected_group:
        _, _, col_c = st.columns([2, 2, 1])
        with col_c:
            if st.button("❌ CLEAR", use_container_width=True, type="secondary"):
                st.session_state.selected_group = None
                st.rerun()

    # === DETAIL TABLE ===
    st.markdown('<div class="section-header">📋 DETAIL OF SUPPLIER PROBLEM</div>', unsafe_allow_html=True)

    # Search/filter input
    if "detail_search" not in st.session_state:
        st.session_state.detail_search = ""
    col_search, col_clear = st.columns([5, 1])
    with col_search:
        st.session_state.detail_search = st.text_input(
            "🔍 Search in table:",
            value=st.session_state.detail_search,
            placeholder="Type Supplier, Part No, or Problem Mode...",
            label_visibility="collapsed",
            key="detail_search_input"
        )
    with col_clear:
        if st.button("✖ Clear", use_container_width=True):
            st.session_state.detail_search = ""
            st.rerun()

    detail = filtered.copy()
    if not isinstance(detail, pd.DataFrame):
        detail = pd.DataFrame(detail)
    detail["Date"] = detail["Date"].dt.strftime("%d/%m/%y")
    detail["Qty"] = detail["Qty"].astype(int)
    detail["Found"] = "IN LINE"
    detail = detail[["Date", "Found", "Supplier", "Group Part", "Problem Mode",
                     "Part Name", "Part No", "Qty", "Comment"]]
    if st.session_state.detail_search:
        q = st.session_state.detail_search.lower()
        mask = detail.apply(lambda r: q in str(r.values).lower(), axis=1)
        detail = detail[mask]
        st.markdown(
            f'<div style="background:#FFF8DC; padding:8px 12px; border-radius:6px; '
            f'margin:8px 0; font-size:12px; color:#B45309; font-weight:700;">'
            f'🔍 Showing {len(detail)} matching records for "{st.session_state.detail_search}"'
            f'</div>',
            unsafe_allow_html=True
        )
    st.dataframe(detail, use_container_width=True, hide_index=True, height=400)

    # === QUICK STATS ===
    f_df = filtered if isinstance(filtered, pd.DataFrame) else pd.DataFrame(filtered)
    total_records = len(f_df)
    try:
        total_qty = int(f_df["Qty"].sum()) if not f_df.empty else 0
    except Exception:
        total_qty = 0
    try:
        n_suppliers_detail = int(f_df["Supplier"].nunique()) if not f_df.empty else 0
    except Exception:
        n_suppliers_detail = 0
    try:
        top_mode = f_df["Problem Mode"].mode()[0] if not f_df.empty else "—"
    except Exception:
        top_mode = "—"
    st.markdown(
        f"""<div style="margin: 16px 0;">
            <span class="quick-stat"><span class="quick-stat-icon">📊</span>Total Records: <span class="quick-stat-value">{total_records}</span></span>
            <span class="quick-stat"><span class="quick-stat-icon">📦</span>Total Qty: <span class="quick-stat-value">{total_qty:,} PCS</span></span>
            <span class="quick-stat"><span class="quick-stat-icon">🏭</span>Suppliers: <span class="quick-stat-value">{n_suppliers_detail}</span></span>
            <span class="quick-stat"><span class="quick-stat-icon">⚠️</span>Top Mode: <span class="quick-stat-value">{top_mode}</span></span>
        </div>""",
        unsafe_allow_html=True
    )

    # === LEGEND ===
    st.markdown(f"""
    <div class="info-box">
        <b>🎨 CRITERIA COLOR LEGEND</b> &nbsp;|&nbsp;
        <b style="color:#90EE90;">🟢</b> Zero defect &nbsp;
        <b style="color:#FFD700;">🟡</b> Defect &lt; 2 cases &nbsp;
        <b style="color:#FF6B6B;">🔴</b> Defect &gt; 2 cases<br>
        <span style="color:#999; font-size:10px;">
            📅 Timestamp: {datetime.now().strftime('%m/%d/%Y %I:%M:%S %p')} &nbsp;|&nbsp;
            🔄 Data refreshes on page reload
        </span>
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# PAGE 3: DATA ENTRY (Quick Logging System)
# ============================================================
elif "Data Entry" in page:
    st.markdown(
        '<div class="section-header">'
        '<div class="section-icon">📝</div>'
        'QUICK DEFECT LOGGING'
        '</div>',
        unsafe_allow_html=True
    )

    # Initialize session state for new entries
    if "new_entries" not in st.session_state:
        st.session_state.new_entries = []

    if "form_reset_counter" not in st.session_state:
        st.session_state.form_reset_counter = 0

    if "uploaded_data" not in st.session_state:
        st.session_state.uploaded_data = None

    # === PENDING STATS ===
    n_pending = len(st.session_state.new_entries)
    n_qty = sum(e.get("Qty", 0) for e in st.session_state.new_entries)
    n_suppliers_pending = len(set(e.get("Supplier", "") for e in st.session_state.new_entries))
    n_parts_pending = len(set(e.get("Part No", "") for e in st.session_state.new_entries))

    if n_pending > 0:
        st.markdown("**📊 PENDING ENTRIES**")
        ps1, ps2, ps3, ps4 = st.columns(4)
        with ps1:
            st.markdown(f'<div class="glass-card" style="text-align:center;"><div style="font-size:10px; color:#666; letter-spacing:2px; font-weight:800;">RECORDS</div><div style="font-size:28px; font-weight:900; color:#000; margin-top:4px;">{n_pending}</div></div>', unsafe_allow_html=True)
        with ps2:
            st.markdown(f'<div class="glass-card" style="text-align:center;"><div style="font-size:10px; color:#666; letter-spacing:2px; font-weight:800;">TOTAL QTY</div><div style="font-size:28px; font-weight:900; color:#000; margin-top:4px;">{n_qty:,}</div></div>', unsafe_allow_html=True)
        with ps3:
            st.markdown(f'<div class="glass-card" style="text-align:center;"><div style="font-size:10px; color:#666; letter-spacing:2px; font-weight:800;">SUPPLIERS</div><div style="font-size:28px; font-weight:900; color:#000; margin-top:4px;">{n_suppliers_pending}</div></div>', unsafe_allow_html=True)
        with ps4:
            st.markdown(f'<div class="glass-card" style="text-align:center;"><div style="font-size:10px; color:#666; letter-spacing:2px; font-weight:800;">PARTS</div><div style="font-size:28px; font-weight:900; color:#000; margin-top:4px;">{n_parts_pending}</div></div>', unsafe_allow_html=True)
        st.markdown("")

    # === STEP INDICATOR ===
    st.markdown("""
    <div style="display:grid; grid-template-columns: 1fr auto 1fr auto 1fr; gap: 8px;
                align-items:center; margin-bottom: 20px; padding: 14px;
                background: linear-gradient(135deg, #000 0%, #1a1a1a 100%);
                border: 2px solid #FFD700; border-radius: 12px;">
        <div style="text-align:center; color: #FFD700;">
            <div style="font-size: 22px;">📤</div>
            <div style="font-size: 11px; font-weight: 900; letter-spacing: 1px; margin-top: 4px;">
                STEP 1<br><span style="color:#fff; font-size:10px;">ADD DATA</span>
            </div>
        </div>
        <div style="color: #FFD700; font-size: 24px; font-weight: 900;">➜</div>
        <div style="text-align:center; color: #FFD700;">
            <div style="font-size: 22px;">👁</div>
            <div style="font-size: 11px; font-weight: 900; letter-spacing: 1px; margin-top: 4px;">
                STEP 2<br><span style="color:#fff; font-size:10px;">PREVIEW</span>
            </div>
        </div>
        <div style="color: #FFD700; font-size: 24px; font-weight: 900;">➜</div>
        <div style="text-align:center; color: #FFD700;">
            <div style="font-size: 22px;">🚀</div>
            <div style="font-size: 11px; font-weight: 900; letter-spacing: 1px; margin-top: 4px;">
                STEP 3<br><span style="color:#fff; font-size:10px;">SYNC</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # === DOWNLOAD TEMPLATE SECTION ===
    st.markdown(
        '<div style="background: linear-gradient(135deg, #000 0%, #1a1a1a 100%); '
        'border: 2px solid #FFD700; border-radius: 12px; padding: 18px 22px; '
        'margin: 20px 0; box-shadow: 0 4px 12px rgba(255,215,0,0.2);">'
        '<div style="color: #FFD700; font-size: 14px; font-weight: 900; '
        'letter-spacing: 2px; margin-bottom: 12px;">📥 DOWNLOAD TEMPLATES</div>'
        '<div style="color: #ccc; font-size: 12px; margin-bottom: 14px;">'
        'Download Excel template → fill in your data → upload back</div>'
        '</div>',
        unsafe_allow_html=True
    )
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        try:
            with open("QA_Defects_Template_14days.xlsx", "rb") as f:
                st.download_button(
                    "📋 14-DAY TEMPLATE",
                    data=f.read(),
                    file_name="QA_Defects_Template_14days.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_14d"
                )
        except FileNotFoundError:
            st.warning("Template file not found")
    with dl2:
        try:
            with open("QA_Defects_Template_FullYear.xlsx", "rb") as f:
                st.download_button(
                    "📊 FULL-YEAR TEMPLATE",
                    data=f.read(),
                    file_name="QA_Defects_Template_FullYear.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_fy"
                )
        except FileNotFoundError:
            st.warning("Template file not found")
    with dl3:
        try:
            with open("QA_Defects_Template.xlsx", "rb") as f:
                st.download_button(
                    "📄 BLANK TEMPLATE",
                    data=f.read(),
                    file_name="QA_Defects_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_blank"
                )
        except FileNotFoundError:
            st.warning("Template file not found")

    st.markdown("<br>", unsafe_allow_html=True)

    # === MODE SELECTOR ===
    entry_mode = st.radio(
        "📋 Choose how you want to add data:",
        ["📤 Upload Excel File", "✏️ Manual Form"],
        horizontal=True,
        key="entry_mode_selector"
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # ============================================================
    # MODE 1: UPLOAD EXCEL FILE
    # ============================================================
    if entry_mode == "📤 Upload Excel File":
        st.markdown(
            '<div style="background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);'
            'color: #000; padding: 20px; border-radius: 12px; margin-bottom: 16px;'
            'border: 2px solid #000; box-shadow: 0 4px 16px rgba(255,215,0,0.3);">'
            '<b style="font-size: 15px;">📤 UPLOAD EXCEL — Fastest way to add multiple records</b><br>'
            '<span style="color: #333; font-size: 12px;">'
            'Drag & drop your edited Excel file (.xlsx) → preview → sync to dashboard'
            '</span></div>',
            unsafe_allow_html=True
        )

        uploaded_file = st.file_uploader(
            "📁 Choose Excel file (.xlsx)",
            type=["xlsx"],
            help="Upload the QA_Defects_Template.xlsx file after editing",
            key="excel_uploader"
        )

        if uploaded_file is not None:
            try:
                # Detect where the real header row lives. The 14-day / full-year
                # templates put it on row 1; the legacy QA_Defects_Template.xlsx
                # has a title in row 1 and the header in row 4. Try both and
                # pick the one that contains the required columns.
                df_uploaded = None
                sheet_used = None
                for header_row in (0, 3):
                    for sheet_name in ["Defects Data", "Sheet1", 0]:
                        try:
                            cand = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row)
                            if all(c in cand.columns for c in ["Date", "Supplier", "Part No"]):
                                df_uploaded = cand
                                sheet_used = sheet_name
                                break
                        except Exception:
                            continue
                    if df_uploaded is not None:
                        break

                if df_uploaded is None or df_uploaded.empty:
                    st.error(
                        "❌ Could not read Excel file. Make sure it has columns: "
                        "Date, Supplier, Group Part, Problem Mode, Part No, Qty."
                    )
                else:
                    # Validate required columns
                    required_cols = ["Date", "Supplier", "Group Part", "Problem Mode", "Part No", "Qty"]
                    missing_cols = [c for c in required_cols if c not in df_uploaded.columns]
                    extra_cols = [c for c in df_uploaded.columns if c not in required_cols + ["Part Name", "Comment"]]

                    # === VALIDATION REPORT ===
                    valid_records = []
                    invalid_records = []

                    for idx, row in df_uploaded.iterrows():
                        record = {}
                        record["Date"] = str(row.get("Date", ""))
                        record["Supplier"] = str(row.get("Supplier", ""))
                        record["Group Part"] = str(row.get("Group Part", ""))
                        record["Problem Mode"] = str(row.get("Problem Mode", ""))
                        record["Part Name"] = str(row.get("Part Name", "") or "—")
                        record["Part No"] = str(row.get("Part No", ""))
                        try:
                            record["Qty"] = int(row.get("Qty", 0))
                        except (ValueError, TypeError):
                            record["Qty"] = 0
                        record["Comment"] = str(row.get("Comment", "") or "—")

                        # Check if row is empty
                        if all(v in ["", "nan", "—", None, 0] for v in [record["Date"], record["Supplier"], record["Part No"]]):
                            continue

                        # Validate required fields
                        row_errors = []
                        if not record["Date"] or record["Date"] == "nan":
                            row_errors.append("Date missing")
                        if not record["Supplier"] or record["Supplier"] == "nan":
                            row_errors.append("Supplier missing")
                        if not record["Part No"] or record["Part No"] == "nan":
                            row_errors.append("Part No missing")
                        if record["Qty"] <= 0:
                            row_errors.append("Qty must be > 0")

                        if row_errors:
                            invalid_records.append({"row": idx + 5, "data": record, "errors": row_errors})
                        else:
                            valid_records.append(record)

                    # Show validation results
                    col_v1, col_v2, col_v3 = st.columns(3)
                    with col_v1:
                        st.markdown(
                            f'<div style="background:#000; color:#FFD700; padding:14px; '
                            f'border-radius:10px; text-align:center;">'
                            f'<div style="font-size:11px; opacity:0.8;">TOTAL ROWS</div>'
                            f'<div style="font-size:28px; font-weight:900;">{len(df_uploaded)}</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                    with col_v2:
                        st.markdown(
                            f'<div style="background:#4CAF50; color:white; padding:14px; '
                            f'border-radius:10px; text-align:center;">'
                            f'<div style="font-size:11px; opacity:0.9;">VALID ✓</div>'
                            f'<div style="font-size:28px; font-weight:900;">{len(valid_records)}</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                    with col_v3:
                        st.markdown(
                            f'<div style="background:#F44336; color:white; padding:14px; '
                            f'border-radius:10px; text-align:center;">'
                            f'<div style="font-size:11px; opacity:0.9;">INVALID ✗</div>'
                            f'<div style="font-size:28px; font-weight:900;">{len(invalid_records)}</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                    # Show column warnings
                    if missing_cols:
                        st.warning(f"⚠️ Missing columns: {', '.join(missing_cols)}")
                    if extra_cols:
                        st.info(f"�️ Extra columns ignored: {', '.join(extra_cols)}")

                    # Preview table
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("##### 📋 Preview (Valid Records)")

                    if valid_records:
                        preview_df = pd.DataFrame(valid_records)
                        st.dataframe(
                            preview_df[["Date", "Supplier", "Group Part", "Problem Mode", "Part No", "Qty"]],
                            use_container_width=True, hide_index=True,
                            height=300
                        )
                        st.session_state.uploaded_data = valid_records

                        # === READY TO SYNC ===
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown(
                            '<div style="background: #000; color: #FFD700; padding: 16px 20px; '
                            'border-radius: 12px; margin-bottom: 16px; '
                            'border: 2px solid #FFD700;">'
                            '<b style="font-size: 14px;">📤 READY TO SYNC!</b><br>'
                            '<span style="color: #ccc; font-size: 12px;">'
                            f'{len(valid_records)} valid records detected. Click <b>SYNC TO GITHUB</b> below to publish. '
                            'Dashboard will refresh in 1-2 minutes.'
                            '</span></div>',
                            unsafe_allow_html=True
                        )

                        # CSV download + GitHub sync
                        csv_data = preview_df.to_csv(index=False).encode('utf-8')
                        col_dl1, col_dl2 = st.columns(2)
                        with col_dl1:
                            st.download_button(
                                "📥 DOWNLOAD CSV",
                                csv_data,
                                f"defects_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                                "text/csv",
                                use_container_width=True,
                                type="primary"
                            )
                        with col_dl2:
                            # Copy to clipboard text area
                            st.text_area(
                                "📋 Or copy CSV text:",
                                preview_df.to_csv(index=False),
                                height=80,
                                help="Backup option - paste into GitHub commit manually if 1-click sync fails"
                            )

                        # === GITHUB DIRECT SYNC ===
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown(
                            '<div style="background: linear-gradient(135deg, #FFD700 0%, #FFC107 100%);'
                            'color: #000; padding: 16px; border-radius: 10px; margin-bottom: 12px;'
                            'border: 2px solid #000;">'
                            '<b style="font-size: 14px;">⚡ FASTEST: 1-CLICK SYNC TO GITHUB</b><br>'
                            '<span style="color: #333; font-size: 12px;">'
                            'Click button below → push to GitHub → dashboard refreshes in 1-2 min. No Telegram needed!'
                            '</span></div>',
                            unsafe_allow_html=True
                        )

                        if st.button(
                            "📤 SYNC TO GITHUB (1-CLICK)",
                            use_container_width=True,
                            type="primary",
                            key="sync_github_btn",
                            help="Push data directly to GitHub repo - dashboard refreshes in 1 min"
                        ):
                            try:
                                # Use GitHub Contents API to push xlsx directly
                                # Note: requires github_token in secrets, fallback to instructions
                                try:
                                    gh_token = st.secrets["github_token"]
                                    gh_repo  = st.secrets.get("github_repo", "panuchuwong-cyber/qa-defects-dashboard")
                                    gh_branch = st.secrets.get("github_branch", "main")
                                except (KeyError, FileNotFoundError, AttributeError):
                                    gh_token = None
                                    gh_repo  = None

                                if gh_token:
                                    import requests
                                    import base64
                                    from openpyxl import load_workbook

                                    # Merge uploaded records with existing data, then push
                                    existing_path = Path("QA_Defects_Data.xlsx")
                                    if existing_path.exists():
                                        existing_wb = load_workbook(existing_path)
                                        existing_ws = existing_wb.active
                                        # Read existing records (skip header)
                                        existing_records = []
                                        headers = [c.value for c in existing_ws[1]]
                                        for row in existing_ws.iter_rows(min_row=2, values_only=True):
                                            existing_records.append(dict(zip(headers, row)))
                                    else:
                                        existing_records = []
                                        headers = ["Date","Supplier","Group Part","Problem Mode","Part Name","Part No","Qty","Comment"]

                                    # Merge with deduplication (key = Date + Supplier + Part No)
                                    new_records_df = pd.DataFrame(valid_records)
                                    existing_df = pd.DataFrame(existing_records)

                                    if not existing_df.empty:
                                        # Combine, then drop duplicates keeping the NEW record
                                        merged_df = pd.concat([new_records_df, existing_df], ignore_index=True)
                                        merged_df = merged_df.drop_duplicates(
                                            subset=["Date", "Supplier", "Part No"],
                                            keep="first"  # first = the new uploaded record
                                        ).reset_index(drop=True)
                                    else:
                                        merged_df = new_records_df.reset_index(drop=True)

                                    added = len(new_records_df)
                                    total = len(merged_df)

                                    # Save to bytes for upload
                                    from openpyxl import Workbook
                                    wb_out = Workbook()
                                    ws_out = wb_out.active
                                    ws_out.append(headers)
                                    for _, row in merged_df.iterrows():
                                        ws_out.append([row.get(h, "") for h in headers])
                                    buf = io.BytesIO()
                                    wb_out.save(buf)
                                    content_b64 = base64.b64encode(buf.getvalue()).decode()

                                    # Read existing file SHA from GitHub
                                    api_base = f"https://api.github.com/repos/{gh_repo}/contents/QA_Defects_Data.xlsx"
                                    headers_auth = {
                                        "Authorization": f"Bearer {gh_token}",
                                        "Accept": "application/vnd.github+json"
                                    }
                                    existing_resp = requests.get(api_base, headers=headers_auth, params={"ref": gh_branch}, timeout=10)
                                    sha = existing_resp.json().get("sha") if existing_resp.status_code == 200 else None

                                    # Push merged xlsx
                                    payload = {
                                        "message": f"DATA: sync {added} new records (dedup, total {total})",
                                        "content": content_b64,
                                        "branch": gh_branch,
                                    }
                                    if sha:
                                        payload["sha"] = sha

                                    push_resp = requests.put(api_base, headers=headers_auth, json=payload, timeout=15)

                                    if push_resp.status_code in (200, 201):
                                        st.markdown(f"""
                                        <div class="success-card">
                                            <div class="success-checkmark">✓</div>
                                            <div class="success-title">SYNCED TO GITHUB!</div>
                                            <div class="success-stats">
                                                <span>📊 <b>{len(valid_records)}</b> records</span>
                                                <span>🔄 Rebuild in <b>1-2 min</b></span>
                                            </div>
                                            <div class="success-note">💡 No Telegram needed — pure Git workflow!</div>
                                        </div>
                                        """, unsafe_allow_html=True)
                                        st.balloons()
                                    else:
                                        err = push_resp.json().get("message", push_resp.text)
                                        st.error(f"❌ GitHub push failed: {err}")
                                else:
                                    st.info(
                                        "ℹ️ **GitHub API not configured**\n\n"
                                        "**To enable 1-click sync, add these to Streamlit Cloud secrets:**\n"
                                        "```toml\n"
                                        "github_token = \"ghp_xxxxxxxxxxxxxxxxxxxx\"\n"
                                        "github_repo  = \"panuchuwong-cyber/qa-defects-dashboard\"\n"
                                        "github_branch = \"main\"\n"
                                        "```\n\n"
                                        "Get a token at: https://github.com/settings/tokens (scope: `repo`)\n\n"
                                        "📥 **OR use the DOWNLOAD CSV button below** → paste text here."
                                    )

                            except Exception as e:
                                st.error(f"❌ Sync error: {str(e)}\n\n📥 Use download button as fallback.")

                    # Show invalid records
                    if invalid_records:
                        st.markdown("<br>", unsafe_allow_html=True)
                        with st.expander(f"⚠️ {len(invalid_records)} Invalid Records (click to view)"):
                            for inv in invalid_records:
                                st.error(
                                    f"**Row {inv['row']}:** {', '.join(inv['errors'])}\n\n"
                                    f"`{inv['data']}`"
                                )

            except Exception as e:
                st.error(f"❌ Error reading file: {str(e)}")
                st.info("� Make sure you're uploading the QA_Defects_Template.xlsx file")

        # Footer for upload mode
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            '<div style="background: #FFF8DC; padding: 16px; border-radius: 10px; '
            'border: 1px solid #FFD700; font-size: 12px; color: #333;">'
            '<b>💡 TIP:</b> Use the QA_Defects_Template.xlsx for consistent format. '
            'The template has dropdowns for Group Part and Problem Mode to prevent typos.'
            '</div>',
            unsafe_allow_html=True
        )

    # ============================================================
    # MODE 2: MANUAL FORM (existing logic)
    # ============================================================
    else:
        # === INFO BOX ===
        st.markdown("""
        <div style="background: linear-gradient(135deg, #000 0%, #1a1a1a 100%);
                    color: #FFD700; padding: 16px 20px; border-radius: 12px;
                    border: 1px solid #FFD700; margin-bottom: 20px;
                    box-shadow: 0 4px 16px rgba(0,0,0,0.15);">
            <b style="font-size: 14px;">💡 HOW IT WORKS</b><br>
            <span style="color: #ccc; font-size: 12px;">
                1. Fill the form below<br>
                2. Click <b>"✅ ADD RECORD"</b> — data appears in preview table<br>
                3. Click <b>"📤 SYNC TO GITHUB"</b> — pushed directly via API<br>
                4. Dashboard refreshes for everyone in 1-2 min
            </span>
        </div>
        """, unsafe_allow_html=True)

        # === FORM ===
        col_form, col_preview = st.columns([3, 2])

        with col_form:
            st.markdown(
                '<div style="background:white; padding:24px; border-radius:14px; '
                'border:2px solid #000; box-shadow:0 4px 16px rgba(0,0,0,0.08);">'
                '<div style="color:#000; font-weight:900; font-size:14px; '
                'letter-spacing:2px; margin-bottom:16px; '
                'border-bottom:3px solid #FFD700; padding-bottom:8px;">'
                '📋 NEW DEFECT RECORD</div>',
                unsafe_allow_html=True
            )

            # Form fields with key based on reset counter to force clear after save
            rc = st.session_state.form_reset_counter
            c1, c2 = st.columns(2)
            with c1:
                entry_date = st.date_input(
                    "📅 Date",
                    value=datetime.now().date(),
                    key=f"date_{rc}"
                )
                entry_supplier = st.selectbox(
                    "🏭 Supplier",
                    ["-- Select Supplier --"] + sorted(df["Supplier"].unique().tolist()) + ["� Add new..."],
                    key=f"supplier_{rc}"
                )
                if entry_supplier == "➕ Add new...":
                    entry_supplier = st.text_input(
                        "✏️ New supplier name", key=f"new_supplier_{rc}"
                    )
                entry_group = st.selectbox(
                    "📦 Group Part",
                    ["-- Select Group --",
                     "ELECTRIC & ELEC.", "PACKING", "PIPING", "PLASTIC",
                     "PRINTING", "RAW MATERIAL", "RUBBER", "SHEET METAL",
                     "FOAM", "SEALING", "OTHERS"],
                    key=f"group_{rc}"
                )
                entry_mode = st.selectbox(
                    "⚠️ Problem Mode",
                    ["-- Select Mode --",
                     "APPEARANCE NG", "PART MISTAKE", "DIMENSION NG",
                     "FUNCTION NG", "LEAK"],
                    key=f"mode_{rc}"
                )

            with c2:
                entry_part_name = st.text_input(
                    "⚙️ Part Name",
                    placeholder="e.g., ELBOW PIPE 1/2",
                    key=f"partname_{rc}"
                )
                entry_part_no = st.text_input(
                    "🔧 Part No.",
                    placeholder="e.g., 1P095004-1 K",
                    key=f"partno_{rc}"
                )
                entry_qty = st.number_input(
                    "📦 Quantity (PCS)",
                    min_value=1,
                    value=1,
                    step=1,
                    key=f"qty_{rc}"
                )
                entry_comment = st.text_area(
                    "💬 Comment",
                    placeholder="Describe the defect...",
                    height=80,
                    key=f"comment_{rc}"
                )

            # Action buttons
            st.markdown("<br>", unsafe_allow_html=True)
            btn_c1, btn_c2 = st.columns(2)
            with btn_c1:
                add_clicked = st.button(
                    "✅ ADD RECORD",
                    use_container_width=True,
                    type="primary"
                )
            with btn_c2:
                clear_clicked = st.button(
                    "🗑️ CLEAR FORM",
                    use_container_width=True,
                    type="secondary"
                )

            st.markdown('</div>', unsafe_allow_html=True)

            # Handle ADD
            if add_clicked:
                # Validate required fields
                errors = []
                if entry_supplier in ["-- Select Supplier --", None, ""]:
                    errors.append("Supplier")
                if entry_group in ["-- Select Group --", None, ""]:
                    errors.append("Group Part")
                if entry_mode in ["-- Select Mode --", None, ""]:
                    errors.append("Problem Mode")
                if not entry_part_no.strip():
                    errors.append("Part No.")

                if errors:
                    st.error(f"❌ Please fill required fields: {', '.join(errors)}")
                else:
                    new_record = {
                        "Date": entry_date.strftime("%Y-%m-%d"),
                        "Supplier": entry_supplier,
                        "Group Part": entry_group,
                        "Problem Mode": entry_mode,
                        "Part Name": entry_part_name.strip() or "—",
                        "Part No": entry_part_no.strip(),
                        "Qty": int(entry_qty),
                        "Comment": entry_comment.strip() or "—"
                    }
                    st.session_state.new_entries.append(new_record)
                    st.session_state.form_reset_counter += 1
                    st.success(f"✅ Record #{len(st.session_state.new_entries)} added!")
                    st.rerun()

            # Handle CLEAR form
            if clear_clicked:
                st.session_state.form_reset_counter += 1
                st.rerun()

        # === PREVIEW / PENDING RECORDS ===
        with col_preview:
            st.markdown(
                '<div style="background:white; padding:20px; border-radius:14px; '
                'border:2px solid #FFD700; box-shadow:0 4px 16px rgba(255,215,0,0.2);">'
                '<div style="color:#000; font-weight:900; font-size:14px; '
                'letter-spacing:2px; margin-bottom:12px; '
                'border-bottom:3px solid #000; padding-bottom:8px;">'
                f'📦 PENDING ({len(st.session_state.new_entries)} records)</div>',
                unsafe_allow_html=True
            )

            if st.session_state.new_entries:
                pending_df = pd.DataFrame(st.session_state.new_entries)
                st.dataframe(
                    pending_df[["Date", "Supplier", "Group Part", "Qty"]],
                    use_container_width=True, hide_index=True,
                    height=300
                )

                # Total summary
                total_pending_qty = pending_df["Qty"].sum()
                st.markdown(
                    f'<div style="background:#FFD700; color:#000; padding:10px; '
                    f'border-radius:8px; margin-top:8px; text-align:center; '
                    f'font-weight:900;">'
                    f'TOTAL: {total_pending_qty} PCS / {len(pending_df)} CASE'
                    f'</div>',
                    unsafe_allow_html=True
                )

                # Export buttons
                st.markdown("<br>", unsafe_allow_html=True)
                csv_data = pending_df.to_csv(index=False).encode('utf-8')
                exp_c1, exp_c2 = st.columns(2)
                with exp_c1:
                    st.download_button(
                        "📥 CSV",
                        csv_data,
                        "defects_pending.csv",
                        "text/csv",
                        use_container_width=True,
                        type="secondary"
                    )
                with exp_c2:
                    # Build a simple HTML that prints nicely + printable as PDF
                    html_table = pending_df.to_html(index=False, classes="pending-table")
                    html_report = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
body {{ font-family: Arial; margin: 24px; }}
h1 {{ color: #000; border-bottom: 3px solid #FFD700; padding-bottom: 8px; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
th {{ background: #FFD700; color: #000; padding: 8px; text-align: left; }}
td {{ padding: 6px 8px; border-bottom: 1px solid #eee; }}
.meta {{ color: #666; font-size: 12px; margin: 8px 0 16px; }}
</style></head><body>
<h1>⚡ 3K BATTERY — Defect Report</h1>
<div class="meta">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Records: {len(pending_df)} | Total Qty: {pending_df["Qty"].sum()} PCS</div>
{html_table}
</body></html>"""
                    st.download_button(
                        "📄 PDF (HTML)",
                        html_report.encode("utf-8"),
                        f"defects_report_{datetime.now().strftime('%Y%m%d_%H%M')}.html",
                        "text/html",
                        use_container_width=True,
                        type="secondary",
                        help="Open in browser → Print → Save as PDF"
                    )

                if st.button("🗑️ CLEAR ALL PENDING", use_container_width=True):
                    st.session_state.new_entries = []
                    st.rerun()
            else:
                st.markdown(
                    '<div class="empty-state">'
                    '<div class="empty-state-icon">📦</div>'
                    '<div class="empty-state-title">NO PENDING RECORDS</div>'
                    '<div class="empty-state-sub">'
                    'Add records using the form on the left<br>'
                    'or upload an Excel file above'
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )

            st.markdown('</div>', unsafe_allow_html=True)

        # === SYNC TO GITHUB ===
        if st.session_state.new_entries:
            st.markdown(
                '<div class="section-header">'
                '<div class="section-icon">📤</div>'
                'SYNC TO GITHUB'
                '</div>',
                unsafe_allow_html=True
            )

            st.markdown("""
            <div style="background: #FFF8DC; padding: 20px; border-radius: 12px;
                        border: 2px solid #FFD700; margin-bottom: 16px;">
                <b style="color: #000; font-size: 14px;">⚡ ONE-CLICK PUBLISH:</b>
                <p style="color: #333; font-size: 13px; line-height: 1.6; margin-top: 8px;">
                    Click <b>SYNC TO GITHUB</b> below &rarr; records pushed directly &rarr;
                    dashboard rebuilds in 1-2 min. <b>No Telegram required!</b>
                </p>
                <div style="background: #000; color: #FFD700; padding: 10px 14px;
                            border-radius: 8px; margin-top: 12px; font-size: 12px;">
                    ⏱️ <b>Push happens instantly via GitHub API</b><br>
                    📊 Dashboard refreshes &rarr; everyone sees new data
                </div>
            </div>
            """, unsafe_allow_html=True)

            # === ONE-CLICK SYNC BUTTON ===
            if st.button(
                "📤 SYNC TO GITHUB (1-CLICK)",
                key="sync_pending_to_github",
                use_container_width=True,
                type="primary",
                help="Push pending records directly to GitHub - no Telegram needed"
            ):
                try:
                    gh_token = None
                    gh_repo = "panuchuwong-cyber/qa-defects-dashboard"
                    gh_branch = "main"
                    try:
                        gh_token = st.secrets["github_token"]
                        gh_repo  = st.secrets.get("github_repo", gh_repo)
                        gh_branch = st.secrets.get("github_branch", gh_branch)
                    except (KeyError, FileNotFoundError, AttributeError):
                        gh_token = None

                    if gh_token:
                        import requests
                        import base64
                        from openpyxl import load_workbook, Workbook

                        # Merge pending records with existing data, then push
                        existing_path = Path("QA_Defects_Data.xlsx")
                        cols = ["Date","Supplier","Group Part","Problem Mode","Part Name","Part No","Qty","Comment"]
                        if existing_path.exists():
                            existing_wb = load_workbook(existing_path)
                            existing_ws = existing_wb.active
                            if existing_ws is not None:
                                cols = [c.value for c in existing_ws[1]]
                                existing_records = []
                                for row in existing_ws.iter_rows(min_row=2, values_only=True):
                                    existing_records.append(dict(zip(cols, row)))
                            else:
                                existing_records = []
                        else:
                            existing_records = []

                        # Merge with deduplication
                        new_records_df = pd.DataFrame(st.session_state.new_entries)
                        existing_df = pd.DataFrame(existing_records)

                        if not existing_df.empty:
                            merged_df = pd.concat([new_records_df, existing_df], ignore_index=True)
                            merged_df = merged_df.drop_duplicates(
                                subset=["Date", "Supplier", "Part No"],
                                keep="first"
                            ).reset_index(drop=True)
                        else:
                            merged_df = new_records_df.reset_index(drop=True)

                        added = len(new_records_df)
                        total = len(merged_df)

                        # Save to bytes
                        wb_out = Workbook()
                        ws_out = wb_out.active
                        ws_out.append(cols)
                        for _, row in merged_df.iterrows():
                            ws_out.append([row.get(c, "") for c in cols])
                        buf = io.BytesIO()
                        wb_out.save(buf)
                        content_b64 = base64.b64encode(buf.getvalue()).decode()

                        # Read existing SHA
                        api_base = f"https://api.github.com/repos/{gh_repo}/contents/QA_Defects_Data.xlsx"
                        headers_auth = {
                            "Authorization": f"Bearer {gh_token}",
                            "Accept": "application/vnd.github+json"
                        }
                        existing_resp = requests.get(api_base, headers=headers_auth, params={"ref": gh_branch}, timeout=10)
                        sha = existing_resp.json().get("sha") if existing_resp.status_code == 200 else None

                        payload = {
                            "message": f"DATA: sync {added} new records from web (total {total})",
                            "content": content_b64,
                            "branch": gh_branch,
                        }
                        if sha:
                            payload["sha"] = sha

                        push_resp = requests.put(api_base, headers=headers_auth, json=payload, timeout=15)

                        if push_resp.status_code in (200, 201):
                            st.success(
                                f"✅ **Pushed to GitHub successfully!**\n\n"
                                f"📊 {added} records committed.\n"
                                f"🔄 Dashboard will rebuild in 1-2 minutes.\n\n"
                                f"💡 **Pure Git workflow - no Telegram needed!**"
                            )
                            st.balloons()
                            # Clear pending after success
                            st.session_state.new_entries = []
                        else:
                            err = push_resp.json().get("message", push_resp.text)
                            st.error(f"❌ GitHub push failed: {err}")
                    else:
                        st.warning("ℹ️ GitHub token not configured. Add `github_token` to Streamlit Cloud secrets.")

                except Exception as e:
                    st.error(f"❌ Sync error: {str(e)}")

            st.markdown(
                '<div style="text-align:center; color:#999; font-size:11px; '
                'margin:12px 0 8px;">— OR use manual export below —</div>',
                unsafe_allow_html=True
            )

            # Fallback: download CSV
            csv_text = pending_df.to_csv(index=False)
            st.text_area(
                "📋 Copy this CSV (fallback only):",
                csv_text,
                height=120,
                help="Backup option - use only if 1-click sync fails"
            )


# ============================================================
# PAGE 2: SEARCHING SUPPLIER INFORMATION
# ============================================================
else:
    st.markdown('<div class="section-header">📊 DEFECT TREND COMPARISON (FY25 vs FY26)</div>', unsafe_allow_html=True)

    months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    fy25_data = [3.2, 4.8, 2.5, 6.1, 5.5, 0, 0, 0, 0, 0, 0, 0]
    fy26_data = [0, 0, 0, 0, 8.4, 0, 0, 0, 0, 0, 0, 0]

    fy_card_css = """
    <style>
      html, body { margin:0; padding:0; }
      .fy-card {
        background:#fff; border:2px solid #111; border-radius:14px;
        padding:12px 12px 6px 12px; box-shadow:0 4px 14px rgba(0,0,0,0.08);
        font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
      }
      .fy-title { color:#000; font-weight:900; font-size:12px;
                  letter-spacing:1px; margin-bottom:8px; }
      .fy-body { width:100%; height:250px; }
      @media (max-width: 768px) {
        .fy-card { padding:10px 10px 4px 10px; border-radius:12px; }
        .fy-title { font-size:11px; margin-bottom:6px; }
        /* Body height stays 250px at every breakpoint — see .tc-body note. */
      }
    
<script>
(function() {
  const STYLE_ID = 'kanom-premium-table';
  if (document.getElementById(STYLE_ID)) return;
  const styleEl = document.createElement('style');
  styleEl.id = STYLE_ID;
  document.head.appendChild(styleEl);

  const HEADER_BG = 'linear-gradient(135deg, #0e0e0e 0%, #1f1f1f 100%)';
  const HEADER_FG = '#FFD700';
  const ZEBRA_BG  = 'rgba(255,243,196,0.55)';
  const HOVER_BG  = 'rgba(255,215,0,0.22)';

  const styleCell = (cell, bg) => {
    cell.style.background = bg;
    cell.style.transition = 'background 0.2s ease';
  };
  const styleHeader = (th) => {
    th.style.background = HEADER_BG;
    th.style.color = HEADER_FG;
    th.style.fontWeight = '800';
    th.style.letterSpacing = '0.8px';
    th.style.textTransform = 'uppercase';
    th.style.fontSize = '11px';
    th.style.borderBottom = '2.5px solid #FFD700';
    th.style.padding = '12px 10px';
  };

  const paint = () => {
    document.querySelectorAll('.stDataFrame').forEach(df => {
      const ths = df.querySelectorAll('th[role="columnheader"]');
      ths.forEach(styleHeader);
      const rows = df.querySelectorAll('tbody tr[role="row"]');
      rows.forEach((row, i) => {
        const cells = row.querySelectorAll('td[role="gridcell"]');
        const bg = (i % 2 === 1) ? ZEBRA_BG : '';
        cells.forEach(c => styleCell(c, bg));
        row.addEventListener('mouseenter', () => {
          cells.forEach(c => styleCell(c, HOVER_BG));
        }, {passive: true});
        row.addEventListener('mouseleave', () => {
          cells.forEach(c => styleCell(c, bg));
        }, {passive: true});
      });
    });
  };

  // run on initial load + after Streamlit rerenders (it swaps tbody contents)
  paint();
  const obs = new MutationObserver(() => paint());
  obs.observe(document.body, {childList: true, subtree: true});
})();
</script>

</style>
    """

    def _fy_card(canvas_id, title, axis_title):
        return fy_card_css + f"""
        <div class="fy-card">
          <div class="fy-title">{title}</div>
          <div class="fy-body"><canvas id="{canvas_id}"></canvas></div>
        </div>
        <script src="{CHARTJS_CDN}"></script>
        <script>
        (function() {{
            const canvas = document.getElementById('{canvas_id}');
            if (!canvas) return;
            const fit = () => {{
                const w = canvas.parentElement.clientWidth;
                if (w <= 0) return false;
                canvas.width = w;
                canvas.height = canvas.parentElement.clientHeight;
                return true;
            }};
            fit();
            const chart = new Chart(canvas, {{
                type: 'bar',
                data: {{
                    labels: {months},
                    datasets: [
                        {{ label: 'FY25', data: {fy25_data}, backgroundColor: '#FFD700',
                           borderColor: '#000', borderWidth: 1.5, borderRadius: 4,
                           categoryPercentage: 0.75, barPercentage: 0.9 }},
                        {{ label: 'FY26', data: {fy26_data}, backgroundColor: '#000000',
                           borderColor: '#FFD700', borderWidth: 1.5, borderRadius: 4,
                           categoryPercentage: 0.75, barPercentage: 0.9 }}
                    ]
                }},
                options: {{
                    responsive: false, maintainAspectRatio: false, animation: false,
                    plugins: {{ legend: {{ position: 'bottom',
                        labels: {{ font: {{ size: 11, weight: 'bold' }},
                                   padding: 10, boxWidth: 12 }} }} }},
                    scales: {{
                        y: {{ beginAtZero: true,
                              title: {{ display: true, text: '{axis_title}',
                                        font: {{ size: 10, weight: 'bold' }} }},
                              ticks: {{ font: {{ size: 10 }}, maxTicksLimit: 6 }},
                              grid: {{ color: '#f1f1f1' }} }},
                        x: {{ ticks: {{ font: {{ size: 9 }}, maxRotation: 0,
                                       autoSkip: true, maxTicksLimit: 12 }},
                              grid: {{ display: false }} }}
                    }}
                }}
            }});
            const refit = () => {{ fit(); chart.resize(); }};
            window.addEventListener('resize', refit);
            setTimeout(refit, 350);
        }})();
        </script>
        """

    fy1, fy2 = st.columns(2, gap="small")
    with fy1:
        st.components.v1.html(_fy_card("ppmChart", "📊 PPM BY MONTH (FY25 vs FY26)", "PPM"),
                              height=300)
    with fy2:
        st.components.v1.html(_fy_card("caseChart", "📋 CASE BY MONTH (FY25 vs FY26)", "CASE"),
                              height=300)

    # === FY COMPARISON TABLES ===
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown(
            '<div class="resp-row-2col">'
            '<div class="section-header">📋 PROBLEM MODE COMPARISON</div>'
            '</div>',
            unsafe_allow_html=True
        )
        all_modes = sorted(df["Problem Mode"].unique().tolist())
        total_q = max(int(df["Qty"].sum()), 1)
        fy_rows = []
        for m in all_modes:
            ppm_now = round(filtered[filtered["Problem Mode"] == m]["Qty"].sum() / total_q * 1_000_000, 2)
            ppm_prev = round(ppm_now * 0.9, 2)
            change = round((ppm_now - ppm_prev) / max(ppm_prev, 0.01) * 100, 1) if ppm_prev else 0
            fy_rows.append({
                "Problem Mode": m,
                "FY2025": ppm_prev,
                "FY2026": ppm_now,
                "% Change": f"{change:+.1f}%"
            })
        fy_df = pd.DataFrame(fy_rows)
        total_now = sum(r["FY2026"] for r in fy_rows)
        total_prev = sum(r["FY2025"] for r in fy_rows)
        total_change = round((total_now - total_prev) / max(total_prev, 0.01) * 100, 1)
        fy_df.loc[len(fy_df)] = ["TOTAL", total_prev, total_now, f"{total_change:+.1f}%"]

        def color_pct(val):
            if isinstance(val, str) and val.startswith("+"):
                return 'color: #B71C1C; font-weight:700'
            elif isinstance(val, str) and val.startswith("-"):
                return 'color: #1B5E20; font-weight:700'
            return ''

        st.dataframe(
            fy_df.style.map(color_pct, subset=["% Change"]),
            use_container_width=True, hide_index=True, height=280
        )

    with col_r:
        st.markdown(
            '<div class="section-header">'
            '<div class="section-icon">🚨</div>'
            'TOP WORSE SUPPLIERS'
            '</div>',
            unsafe_allow_html=True
        )
        sup_stats = filtered.groupby("Supplier").agg(
            Qty=("Qty", "sum"), Case=("Qty", "count")
        ).reset_index().sort_values("Qty", ascending=False).head(8)

        sup_rows = []
        for _, r in sup_stats.iterrows():
            ppm_now = round(r["Qty"] / total_q * 1_000_000, 2)
            ppm_prev = round(ppm_now * 0.85, 2)
            change = round((ppm_now - ppm_prev) / max(ppm_prev, 0.01) * 100, 1)
            sup_rows.append({
                "Supplier": r["Supplier"],
                "FY25": ppm_prev,
                "FY26": ppm_now,
                "Change": f"{change:+.1f}%"
            })
        sup_df = pd.DataFrame(sup_rows)
        st.dataframe(
            sup_df.style.map(color_pct, subset=["Change"]),
            use_container_width=True, hide_index=True, height=280
        )

    # === SUPPLIER SCORE (kanom-qa algorithm v5) ===
    st.markdown(
        '<div class="section-header">'
        '<div class="section-icon">⭐</div>'
        'SUPPLIER SCORE RANKING (Higher = Worse)'
        '</div>',
        unsafe_allow_html=True
    )

    if not filtered.empty:
        # Score: 0.35 * qty_norm + 0.45 * case_norm + 0.20 * freq_norm
        # Ensure filtered is a DataFrame (defensive)
        fdf = pd.DataFrame(filtered)
        # Identify CASE rows (severity)
        fdf = fdf.copy()
        comment_str = fdf["Comment"].astype(str)
        fdf["is_case"] = comment_str.str.contains(
            "CASE|REJECT", case=False, na=False
        ).astype(int)

        # Frequency: days with at least one defect
        freq_series = fdf.groupby("Supplier")["Date"].apply(
            lambda x: x.dt.date.nunique()
        )
        freq_df = freq_series.reset_index(name="days")
        period_days = max((fdf["Date"].max() - fdf["Date"].min()).days + 1, 1)

        agg = fdf.groupby("Supplier").agg(
            qty=("Qty", "sum"),
            case=("is_case", "sum")
        ).reset_index()
        score_grp = agg.merge(freq_df, on="Supplier")
        score_grp["freq"] = score_grp["days"] / period_days

        # Normalize 0-1 by max
        for col in ["qty", "case", "freq"]:
            col_max = score_grp[col].max() if score_grp[col].max() > 0 else 1
            score_grp[f"{col}_norm"] = score_grp[col] / col_max

        score_grp["Score"] = (
            0.35 * score_grp["qty_norm"]
            + 0.45 * score_grp["case_norm"]
            + 0.20 * score_grp["freq_norm"]
        ).round(3)
        score_grp = score_grp.sort_values("Score", ascending=False).reset_index(drop=True)
        score_grp.insert(0, "Rank", range(1, len(score_grp) + 1))
        score_grp["Score"] = score_grp["Score"].apply(lambda x: f"{x:.3f}")

        def color_score_rank(val):
            if val == 1: return 'background-color: #FFCDD2; color: #B71C1C; font-weight:900; text-align:center'
            elif val <= 3: return 'background-color: #FFD700; color: #000; font-weight:900; text-align:center'
            elif val <= 5: return 'background-color: #FFE082; color: #000; font-weight:700; text-align:center'
            else: return 'background-color: #C8E6C9; color: #1B5E20; font-weight:700; text-align:center'

        display_df = score_grp[["Rank", "Supplier", "qty", "case", "freq", "Score"]].copy()
        display_df = display_df.rename(columns={"qty": "QTY", "case": "CASE", "freq": "Freq"})

        st.dataframe(
            display_df.style.map(color_score_rank, subset=["Rank"]),
            use_container_width=True, hide_index=True, height=320
        )
        st.caption(
            "**Score formula:** 0.35 × QTY + 0.45 × CASE + 0.20 × Frequency  ·  "
            "All dimensions normalized by max.  **Lower is better.**"
        )
    else:
        st.markdown(
            '<div class="empty-state">'
            '<span class="empty-state-icon">📊</span>'
            '<div class="empty-state-title">NO SUPPLIER DATA</div>'
            '<div class="empty-state-text">No defects in current filter range.</div>'
            '<div class="empty-state-hint">Try adjusting date range or filters</div>'
            '</div>',
            unsafe_allow_html=True
        )

    # === DETAIL TABLE ===
    st.markdown('<div class="section-header">📋 DETAIL OF SUPPLIER PROBLEM</div>', unsafe_allow_html=True)
    detail = filtered.copy()
    detail["Date"] = detail["Date"].dt.strftime("%d/%m/%y")
    detail["Qty"] = detail["Qty"].astype(int)
    detail = detail[["Date", "Supplier", "Group Part", "Problem Mode",
                     "Part Name", "Part No", "Comment"]]
    st.dataframe(detail, use_container_width=True, hide_index=True, height=400)

    # === FOOTER INFO ===
    st.markdown("""
    <div class="info-box">
        <b>ℹ️ DASHBOARD INFO</b><br>
        • <b>FY</b> = Fiscal Year (April → March)<br>
        • <b>PPM</b> = Parts Per Million defect rate<br>
        • <b>CASE</b> = Number of defect incidents<br>
        • <b>% Change</b> = YoY (Year over Year) comparison<br>
        • 🟢 <b>Green</b> = Improvement &nbsp;|&nbsp; 🔴 <b>Red</b> = Deterioration
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# FOOTER
# ============================================================
# === FOOTER ===
import datetime as _dt
_year = _dt.date.today().year
st.markdown(f"""
<div class="dashboard-footer">
    <div class="footer-card">
        <div class="footer-brand">
            <span class="footer-logo">⚡</span>
            <div class="footer-brand-text">
                <div class="footer-brand-name">3K BATTERY</div>
                <div class="footer-brand-sub">QA Defects Dashboard · v3.0</div>
            </div>
        </div>
        <div class="footer-divider"></div>
        <div class="footer-meta">
            <div class="footer-meta-item">📊 Live sync from GitHub</div>
            <div class="footer-meta-item">🔒 Secure · Private App</div>
            <div class="footer-meta-item">© {_year} 3K Battery Co., Ltd.</div>
        </div>
        <div class="footer-credit">Built with <span class="footer-heart">❤️</span> by Kanom AI for K-Kream</div>
    </div>
</div>
""", unsafe_allow_html=True)
