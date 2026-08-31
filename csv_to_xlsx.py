"""Convert CSV to Excel (.xlsx) with formatting - matching Daikin style"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

csv_path = "/Users/panuchuwong/Desktop/Kanom/QA_Defects_Dashboard/QA_Defects_Data.csv"
xlsx_path = "/Users/panuchuwong/Desktop/Kanom/QA_Defects_Dashboard/QA_Defects_Data.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Defects Data"

# Read CSV
with open(csv_path, "r", encoding="utf-8-sig") as f:
    reader = csv.reader(f)
    rows = list(reader)

# Write data
for row_idx, row in enumerate(rows, start=1):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)

# Format header (Yellow & Black theme)
header_font = Font(bold=True, color="000000", size=11)
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold yellow
header_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

for col in range(1, len(rows[0]) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Format data rows
data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
for row in range(2, len(rows) + 1):
    for col in range(1, len(rows[0]) + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = data_align
        cell.border = thin_border
        # Alternating row colors (light yellow/white)
        if row % 2 == 0:
            cell.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")

# Auto-width columns (Date | Supplier | Group | Mode | Part Name | Part No | Qty | Comment)
col_widths = [12, 20, 18, 18, 22, 18, 8, 40]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Freeze header
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Metric"
ws2["B1"] = "Value"
for cell in ["A1", "B1"]:
    ws2[cell].font = header_font
    ws2[cell].fill = header_fill
    ws2[cell].alignment = header_align
    ws2[cell].border = thin_border

# Calculate stats (Qty is column index 6 = 0-indexed)
total_qty = sum(int(r[6]) for r in rows[1:])
total_case = len(rows) - 1
unique_suppliers = len(set(r[1] for r in rows[1:]))
unique_parts = len(set(r[5] for r in rows[1:]))

ws2["A2"] = "Total QTY (PCS)"
ws2["B2"] = total_qty
ws2["A3"] = "Total CASE"
ws2["B3"] = total_case
ws2["A4"] = "Unique Suppliers"
ws2["B4"] = unique_suppliers
ws2["A5"] = "Unique Parts"
ws2["B5"] = unique_parts
ws2["A6"] = "Date Range"
ws2["B6"] = "31/07/2026 - 13/08/2026"
ws2["A7"] = "Generated"
ws2["B7"] = "Kanom AI - Aug 2026"

for cell_addr in ["A2", "A3", "A4", "A5", "A6", "A7"]:
    ws2[cell_addr].font = Font(bold=True)
    ws2[cell_addr].border = thin_border
for cell_addr in ["B2", "B3", "B4", "B5", "B6", "B7"]:
    ws2[cell_addr].border = thin_border
    ws2[cell_addr].alignment = Alignment(horizontal="center")

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 30

wb.save(xlsx_path)
print(f"✅ Excel saved: {xlsx_path}")
print(f"   Total QTY: {total_qty}")
print(f"   Total CASE: {total_case}")
print(f"   Unique Suppliers: {unique_suppliers}")
